#!/usr/bin/env python3
"""Weekly Report Formatter v9.50 — Adds Opinionn Reputation block at E34:K43 from Review Summary PDF"""
from flask import Flask, request, send_file, render_template_string, jsonify
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, re, os, traceback
from datetime import datetime
from collections import OrderedDict

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

GREEN='FF7AD694'; GRAY_HDR='FFBFBFBF'; GRAY_AR='FFD9D9D9'
KG_RED='FFF28E86'; KG_YEL='FFFDD868'; KG_BLUE='FF8CB5F9'
WHITE='FFFFFFFF'; BLACK='FF000000'; DARKGRAY='FF505050'; RED_FONT='FFFF0000'
BLUE_IN='FF8CB5F9'

PROPERTY_UNITS = {
    'Village at Madrone (fka Village at Morgan Hill) (x93)': 249,
    'Village at First': 120,
    'Village at Santa Teresa': 100,
}

def get_total_units(prop):
    for key, val in PROPERTY_UNITS.items():
        if key.lower() in prop.lower() or prop.lower() in key.lower():
            return val
    return 249

def gfill(h): return PatternFill(fill_type='solid', fgColor=h)
def gfont(bold=False,sz=9,color='FF000000'): return Font(name='Calibri',size=sz,bold=bold,color=color)
def galign(h='left',v='center',wrap=False): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
T=Side(style='thin',color='FF000000'); TG=Side(style='thin',color='FFCCCCCC')
def bblack(): return Border(top=T,bottom=T,left=T,right=T)
def bgray():  return Border(top=TG,bottom=TG,left=TG,right=TG)

def parse_ua(ws):
    out=[]; status='Occupied'
    rows=list(ws.iter_rows(values_only=True))
    has_kg=False; unit_col=0; unit_pat=r'^\d{3,5}$'
    for row in rows:
        if not row: continue
        v0=str(row[0] or '').strip(); v1=str(row[1] or '').strip() if len(row)>1 else ''
        if re.match(r'^\d{2}-\d{3}',v0): unit_col=0; unit_pat=r'^\d{2}-\d{3}'; break
        if re.match(r'^\d{2}-\d{3}',v1): unit_col=1; unit_pat=r'^\d{2}-\d{3}'; break
        if re.match(r'^\d{3,5}$',v0): unit_col=0; unit_pat=r'^\d{3,5}$'; break
    for row in rows[:10]:
        if any('kg' in str(v or '').lower() for v in row):
            has_kg=True; break
    for row in rows:
        if not row or all(v is None or str(v).strip()=='' for v in row): continue
        c0=str(row[0] or '').strip()
        c1=str(row[1] or '').strip() if len(row)>1 else ''
        hdr=c0 if c0 else c1
        lo=hdr.lower()
        if any(x in lo for x in ['- vacant','- notice','- occupied','- past','- current']):
            if   'occupied' in lo or ('current' in lo and 'notice' not in lo): status='Occupied'
            elif 'vacant' in lo: status='Vacant'
            elif 'notice' in lo: status='Notice'
            continue
        if re.match(r'^(Unit Availability|Showing|Group|As Of|Total|Property)',hdr,re.I): continue
        if hdr in ('Unit',): continue
        if unit_col==1 and str(row[0] or '').strip() in ('Vacant','Notice','Occupied'):
            status=str(row[0]).strip()
        unit=str(row[unit_col] or '').strip() if len(row)>unit_col else ''
        if not re.match(unit_pat, unit): continue
        o=unit_col
        if has_kg:
            out.append({'status':status,'unit':unit,
                'res_id':str(row[o+1] or '').strip() if len(row)>o+1 else '',
                'name':str(row[o+2] or '').strip() if len(row)>o+2 else '',
                'kg_app':row[o+3] if len(row)>o+3 else None,
                'kg_pend':row[o+4] if len(row)>o+4 else None,
                'site_pend':row[o+5] if len(row)>o+5 else None,
                'res_rent':row[o+6] if len(row)>o+6 else None,
                'unit_rent':row[o+7] if len(row)>o+7 else None,
                'res_dep':row[o+8] if len(row)>o+8 else None,
                'unit_dep':row[o+9] if len(row)>o+9 else None,
                'yardi_st':str(row[o+10] or '').strip() if len(row)>o+10 else '',
                'days':row[o+11] if len(row)>o+11 else None,
                'make_rdy':row[o+12] if len(row)>o+12 else None,
                'move_in':row[o+13] if len(row)>o+13 else None,
                'hold':str(row[o+14] or '').strip() if len(row)>o+14 else '',
                'hold_until':row[o+15] if len(row)>o+15 else None,
                'notice':row[o+16] if len(row)>o+16 else None,
                'move_out':row[o+17] if len(row)>o+17 else None,
                'lease_sgn':row[o+18] if len(row)>o+18 else None,
                'lease_from':row[o+19] if len(row)>o+19 else None,
                'lease_to':row[o+20] if len(row)>o+20 else None,
                'has_kg':True})
        else:
            out.append({'status':status,'unit':unit,
                'res_id':str(row[o+1] or '').strip() if len(row)>o+1 else '',
                'name':str(row[o+2] or '').strip() if len(row)>o+2 else '',
                'kg_app':None,'kg_pend':None,'site_pend':None,
                'res_rent':row[o+3] if len(row)>o+3 else None,
                'unit_rent':row[o+4] if len(row)>o+4 else None,
                'res_dep':row[o+5] if len(row)>o+5 else None,
                'unit_dep':row[o+6] if len(row)>o+6 else None,
                'yardi_st':str(row[o+7] or '').strip() if len(row)>o+7 else '',
                'days':row[o+8] if len(row)>o+8 else None,
                'make_rdy':row[o+9] if len(row)>o+9 else None,
                'move_in':row[o+10] if len(row)>o+10 else None,
                'hold':str(row[o+11] or '').strip() if len(row)>o+11 else '',
                'hold_until':row[o+12] if len(row)>o+12 else None,
                'notice':row[o+13] if len(row)>o+13 else None,
                'move_out':row[o+14] if len(row)>o+14 else None,
                'lease_sgn':row[o+15] if len(row)>o+15 else None,
                'lease_from':row[o+16] if len(row)>o+16 else None,
                'lease_to':row[o+17] if len(row)>o+17 else None,
                'has_kg':False})
    return out

def fmt_ua(wb_out, raw_bytes, date, prop):
    wb_r=openpyxl.load_workbook(io.BytesIO(raw_bytes),data_only=True,keep_vba=False,read_only=True)
    data=parse_ua(wb_r.active)
    wb_r.close()
    V=[r for r in data if r['status']=='Vacant']
    N=[r for r in data if r['status']=='Notice']
    O=[r for r in data if r['status']=='Occupied']
    tab=f'Unit Availability {date}'
    if tab in wb_out.sheetnames: del wb_out[tab]
    ws=wb_out.create_sheet(tab); MC=21
    for ti,(text,bold) in enumerate(zip(
        ['Unit Availability Details ',prop,f'As Of: {date}','Showing Pre-Leased: Yes','Showing Occupied: Yes'],
        [True,False,True,False,False])):
        r=ti+1
        for c in range(1,MC+1): ws.cell(r,c).fill=gfill(GREEN); ws.cell(r,c).font=gfont(bold=bold); ws.cell(r,c).alignment=Alignment(horizontal='left',vertical='center',wrap_text=False)
        ws.cell(r,1).value=text
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=MC)
    h6=[None,'Unit','Resident','Name','KG Approved','KG Pend','Site Pending','Resident','Unit','Resident','Unit','Status','Days','Make','Move','Hold','Notice','Move','Lease','Lease','Lease']
    h7=[' ',' ',' ',' ',' ',' ',' ','Rent','Rent','Deposit','Deposit',' ','Vacant','Ready','In',' ',' ','Out','Sign','From','To']
    for c in range(1,MC+1):
        for r,hdr in [(6,h6),(7,h7)]:
            cell=ws.cell(r,c); cell.value=hdr[c-1] if hdr[c-1] is not None else ''
            cell.font=gfont(bold=True); cell.fill=gfill(GRAY_HDR)
            cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=False)
    DATE_COLS={14,15,17,18,19,20,21}
    rn=8; blank=False
    for p in V+N+O:
        isVN=p['status'] in ('Vacant','Notice')
        if p['status']=='Occupied' and not blank:
            for c in range(1,MC+1): ws.cell(rn,c).fill=gfill(WHITE)
            ws.row_dimensions[rn].height=15.0; rn+=1; blank=True
        def sc(col,val,bg=WHITE,h='left',_rn=rn):
            cell=ws.cell(_rn,col); cell.value=val; cell.font=gfont(); cell.fill=gfill(bg)
            cell.alignment=Alignment(horizontal=h,vertical='center',wrap_text=False)
            if col in DATE_COLS and val is not None:
                cell.number_format='MM/DD/YY'
        sc(1,p['status']); sc(2,p['unit']); sc(3,p['res_id'] or None)
        sc(4,None if p['status']=='Vacant' else (p['name'] or None))
        sc(5,None,KG_RED if isVN else WHITE); sc(6,None,KG_YEL if isVN else WHITE); sc(7,None,KG_BLUE if isVN else WHITE)
        sc(8,p['res_rent'] if p['res_rent'] is not None else 0,h='right')
        sc(9,p['unit_rent'] if p['unit_rent'] is not None else 0,h='right')
        sc(10,p['res_dep'] if p['res_dep'] is not None else 0,h='right')
        sc(11,p['unit_dep'] if p['unit_dep'] is not None else 0,h='right')
        sc(12,p['yardi_st'] or None); sc(13,p['days'] if p['days'] is not None else None,h='right')
        sc(14,p['make_rdy'] if p['make_rdy'] is not None else None,h='center')
        sc(15,p['move_in'] if p['move_in'] is not None else None,h='center')
        sc(16,p['hold'] or None)
        sc(17,p['hold_until'] if p['hold_until'] is not None else None,h='center')
        sc(18,p['move_out'] if p['move_out'] is not None else None,h='center')
        sc(19,p['lease_sgn'] if p['lease_sgn'] is not None else None,h='center')
        sc(20,p['lease_from'] if p['lease_from'] is not None else None,h='center')
        sc(21,p['lease_to'] if p['lease_to'] is not None else None,h='center')
        ws.row_dimensions[rn].height=15.0; rn+=1
    for col,w in {'A':14.28,'B':11.42,'C':21.42,'D':15,'E':12,'F':10,'G':12,'H':10,'I':11.42,'J':10,'K':10,'L':7.14,'M':11.42,'N':10,'O':10,'P':8,'Q':10,'R':10,'S':10,'T':10,'U':10}.items():
        ws.column_dimensions[col].width=w
    ws.row_dimensions[1].height=15.75
    for r in range(2,8): ws.row_dimensions[r].height=15.0
    ws.freeze_panes='A10'
    return ws, len(V), len(N), len(O)

def get_notes(wb, prefix):
    notes={}
    tabs=[n for n in wb.sheetnames if prefix.lower() in n.lower()]
    if not tabs: return notes
    rows=list(wb[tabs[-1]].iter_rows(values_only=True))
    hi=next((i for i,r in enumerate(rows[:10]) if any(str(c or '').lower()=='resident' for c in r)),-1)
    if hi==-1: return notes
    hdr=rows[hi]; ri=next((i for i,c in enumerate(hdr) if str(c or '').lower()=='resident'),-1); ni=len(hdr)-1
    for row in rows[hi+1:]:
        id_=str(row[ri] or '').strip() if ri>=0 else ''; note=str(row[ni] or '').strip() if len(row)>ni else ''
        if id_ and note: notes[id_]=note
    return notes

def fmt_ar(wb_out, raw_bytes, date, prev_notes, is_sub, rent_lookup=None):
    """rent_lookup: optional dict mapping unit_number -> tenant_rent (float).
    When provided (Tenant AR only, not SUB AR), rows where any past-due bucket
    (F=0-30, G=31-60, H=61-90, I=Over 90) >= the tenant's full rent get red highlight."""
    wb_r=openpyxl.load_workbook(io.BytesIO(raw_bytes),data_only=True,keep_vba=False,read_only=True)
    rr=list(wb_r.active.iter_rows(values_only=True)); wb_r.close()
    NL='Comments'; tc=BLACK if is_sub else DARKGRAY
    tab=f'{"SUB AR" if is_sub else "Tenant AR"} {date}'
    if tab in wb_out.sheetnames: del wb_out[tab]
    ws=wb_out.create_sheet(tab); MC=13
    for ti in range(3):
        for c in range(1,MC+1):
            ws.cell(ti+1,c).fill=gfill(GREEN); ws.cell(ti+1,c).font=gfont(color=tc)
            ws.cell(ti+1,c).alignment=Alignment(horizontal='center',vertical='center',wrap_text=False)
        ws.cell(ti+1,1).value=str(rr[ti][0] if ti<len(rr) and rr[ti] else '')
    h4=['','','','','Total','','','','','','','','']
    h5=['','','','','Unpaid','0-30','31-60','61-90','Over 90','','','','']
    h6=['Unit','Resident','Status','Name','Charges','days','days','days','days','Prepays','Suspense','Balance','Comments']
    for c in range(1,MC+1):
        ra=5<=c<=12
        for ri,hdr in [(4,h4),(5,h5),(6,h6)]:
            cell=ws.cell(ri,c); cell.value=hdr[c-1]; cell.font=gfont(bold=True)
            cell.fill=gfill(GRAY_AR)
            cell.alignment=Alignment(horizontal='right' if ra else ('center' if c==13 else 'left'),vertical='center',wrap_text=False)
    for r in range(1,4): ws.cell(r,13).fill=gfill(GREEN); ws.cell(r,13).font=gfont(color=tc)
    
    # Add legend for red highlighting (Tenant AR only, not SUB AR)
    if not is_sub:
        ws.merge_cells('N1:N3')
        legend_cell = ws.cell(1, 14)
        legend_cell.value = "Red = tenants who have unpaid full rent balance"
        legend_cell.fill = gfill('FFFFC7CE')  # light red
        legend_cell.font = gfont(bold=True, color=BLACK)
        legend_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # Border the legend
        for r in range(1, 4):
            ws.cell(r, 14).border = bblack()
            ws.cell(r, 14).fill = gfill('FFFFC7CE')
    hi=next((i for i,r in enumerate(rr[:10]) if r and any(str(c or '').lower()=='unit' for c in r) and any(str(c or '').lower()=='resident' for c in r)),5)
    ev,cu,no,cr=[],[],[],[]
    for row in rr[hi+1:]:
        if not row or all(c is None or c=='' for c in row): continue
        st=str(row[2] or '').strip().lower(); u=str(row[0] or '').strip()
        if re.search(r'subtotal|village at|^total$',u,re.I): continue
        if not u or not re.match(r'^\d{2}',u): continue
        charges=row[4]; prepays=row[9] if len(row)>9 else None
        try: charge_val=float(str(charges or 0).replace(',',''))
        except: charge_val=0
        try: prepay_val=float(str(prepays or 0).replace(',',''))
        except: prepay_val=0
        if charge_val < 0 or prepay_val < 0:
            cr.append(row)
        elif st in ('eviction','past'): ev.append(row)
        elif st=='notice': no.append(row)
        else: cu.append(row)
    def sk(r):
        try: return -(float(str(r[4] or 0).replace(',','')))
        except: return 0
    ev.sort(key=sk); cu.sort(key=sk); no.sort(key=sk)
    red_count = [0]  # mutable container so closure can update
    def write_row(rn, row, rc):
        rid=str(row[1] or '').strip(); note=prev_notes.get(rid,'')
        # Determine if this row should be red-highlighted:
        # ONLY for Tenant AR (not SUB AR), and only when rent_lookup is available
        # Logic: any past-due bucket (F/G/H/I = cols 6/7/8/9) >= this tenant's full rent
        has_balance = False
        if not is_sub and rent_lookup:
            unit = str(row[0] or '').strip()
            full_rent = rent_lookup.get(unit, 0)
            if full_rent > 0:
                try:
                    f_val = float(str(row[5] or 0).replace(',',''))   # 0-30 days
                    g_val = float(str(row[6] or 0).replace(',',''))   # 31-60 days
                    h_val = float(str(row[7] or 0).replace(',',''))   # 61-90 days
                    i_val = float(str(row[8] or 0).replace(',',''))   # Over 90 days
                    if max(f_val, g_val, h_val, i_val) >= full_rent:
                        has_balance = True
                        red_count[0] += 1
                except: pass
        row_fill = 'FFFFC7CE' if has_balance else 'FFFFFFFF'  # light red if owes full rent; white otherwise
        for c in range(1,13):
            v=row[c-1]; sv=str(v if v is not None else '').strip()
            try: num=float(sv.replace(',',''))
            except: num=None
            isn=5<=c<=12 and num is not None and sv!=''
            cell=ws.cell(rn,c); cell.value=num if isn else (sv or None)
            cell.font=gfont(color=rc); cell.fill=gfill(row_fill)
            cell.alignment=Alignment(horizontal='right' if c>=5 else 'left',vertical='center',wrap_text=False)
            if isn: cell.number_format='#,##0.00'
        nc=ws.cell(rn,13); nc.value=None; nc.font=gfont(color=BLACK)
        nc.fill=gfill(row_fill); nc.alignment=Alignment(horizontal='center',vertical='center',wrap_text=False)
        nc.number_format='@'
    data_start=7; rn=data_start
    for row in ev+cu+no:
        st=str(row[2] or '').strip().lower()
        rc=RED_FONT if (not is_sub and st in ('notice','eviction','past')) else BLACK
        write_row(rn,row,rc); rn+=1
    tb=bblack()
    pos_end=rn-1
    def write_total(rn, start, end, label='Total'):
        ws.cell(rn,1).value=label; ws.cell(rn,1).font=gfont(bold=True); ws.cell(rn,1).fill=gfill(WHITE); ws.cell(rn,1).border=tb
        for c in range(5,13):
            cell=ws.cell(rn,c)
            cell.value=f'=SUM({get_column_letter(c)}{start}:{get_column_letter(c)}{end})'
            cell.font=gfont(bold=True); cell.fill=gfill(WHITE); cell.border=tb; cell.number_format='#,##0.00'
        for c in [2,3,4,13]: ws.cell(rn,c).fill=gfill(WHITE); ws.cell(rn,c).font=gfont(bold=True); ws.cell(rn,c).border=tb
    write_total(rn,data_start,pos_end); rn+=1
    if cr:
        rn+=1
        for c in range(1,MC+1):
            cell=ws.cell(rn,c); cell.fill=gfill(GRAY_AR); cell.font=gfont(bold=True)
            cell.alignment=Alignment(horizontal='left',vertical='center',wrap_text=False)
        ws.cell(rn,1).value='Credits / Prepays'; ws.cell(rn,4).value='Name'
        ws.cell(rn,5).value='Charges'; ws.cell(rn,10).value='Prepays'; ws.cell(rn,11).value='Suspense'; ws.cell(rn,12).value='Balance'
        for c in [5,10,11,12]: ws.cell(rn,c).alignment=Alignment(horizontal='right',vertical='center',wrap_text=False)
        rn+=1
        cr_start=rn
        for row in cr:
            write_row(rn,row,BLACK); rn+=1
        cr_end=rn-1
        write_total(rn,cr_start,cr_end,'Credits Total'); rn+=1
    for i,w in enumerate([9,13,10,24,12,10,10,10,10,10,10,12,38],1): ws.column_dimensions[get_column_letter(i)].width=w
    if not is_sub:
        ws.column_dimensions['N'].width = 22
    ws.freeze_panes='A7'
    return ws, len(ev), len(cu), len(no), pos_end, data_start, red_count[0]

def fmt_rr(wb_out, raw_bytes, date, prop):
    wb_r=openpyxl.load_workbook(io.BytesIO(raw_bytes),data_only=True,keep_vba=False,read_only=True)
    rr=list(wb_r.active.iter_rows(values_only=True)); wb_r.close()
    tab=f'Rent Roll {date}'
    if tab in wb_out.sheetnames: del wb_out[tab]
    ws=wb_out.create_sheet(tab)

    import datetime as _dt

    MC=15
    title1=str(rr[0][0] or 'FPI Rent Roll') if rr else 'FPI Rent Roll'
    title2=str(rr[1][0] or prop) if len(rr)>1 else prop
    title3=str(rr[2][0] or f'As of Date: {date}') if len(rr)>2 else f'As of Date: {date}'
    title4=str(rr[3][0] or '') if len(rr)>3 else ''
    for r,txt,bold in [(1,title1,True),(2,title2,False),(3,title3,True),(4,title4,False)]:
        for c in range(1,MC+1):
            ws.cell(r,c).fill=gfill(GREEN)
            ws.cell(r,c).font=gfont(bold=bold)
            ws.cell(r,c).alignment=Alignment(horizontal='center',vertical='center',wrap_text=False)
        ws.cell(r,1).value=txt
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=MC)

    h5=['Unit','Unit Type','Unit Set Aside','Resident Name','Sq Ft',
        'Market Rent','Loss/Gain\nto Lease','Sub Rent','Tenant Rent',
        'Lease Rent','Vacancy','Deposit','Move In','Lease From','Lease To']
    for c,h in enumerate(h5,1):
        cell=ws.cell(5,c); cell.value=h; cell.font=gfont(bold=True)
        cell.fill=gfill(GRAY_HDR)
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws.row_dimensions[5].height=28

    hi=next((i for i,r in enumerate(rr[:10]) if r and any(str(c or '').lower()=='unit' for c in r) and any('type' in str(c or '').lower() for c in r)),4)

    rr_offset=0; unit_pat=r'^\d{2}-\d{3}'
    for row in rr[hi+1:]:
        if not row: continue
        v0=str(row[0] or '').strip()
        v1=str(row[1] or '').strip() if len(row)>1 else ''
        if re.match(r'^\d{2}-\d{3}',v0): rr_offset=0; unit_pat=r'^\d{2}-\d{3}'; break
        if re.match(r'^\d{2}-\d{3}',v1): rr_offset=1; unit_pat=r'^\d{2}-\d{3}'; break
        if re.match(r'^\d{3,5}$',v0): rr_offset=0; unit_pat=r'^\d{3,5}'; break
        if re.match(r'^\d{3,5}$',v1): rr_offset=1; unit_pat=r'^\d{3,5}'; break

    is_vaf_format = False
    is_st_format = False
    for row in rr[hi+1:]:
        if not row: continue
        unit_v = str(row[rr_offset] or '').strip()
        if not re.match(unit_pat, unit_v): continue
        col2_v = str(row[rr_offset+2] or '').strip()
        if re.match(r'^\d{3,5}$', col2_v):
            is_vaf_format = True
        break

    V,O=[],[]
    for row in rr[hi+1:]:
        if not row or all(c is None or c=='' for c in row): continue
        if len(row) <= rr_offset: continue
        unit=str(row[rr_offset] or '').strip()
        if not re.match(unit_pat,unit): continue
        o=rr_offset
        name_idx = o+4 if is_vaf_format else o+2
        rname=str(row[name_idx] or '').strip() if len(row)>name_idx else ''
        if rname.strip().upper() in ('VACANT',' VACANT') or not rname.strip(): V.append((row,o))
        else: O.append((row,o))
    V.sort(key=lambda x:str(x[0][x[1]] or '').strip())
    O.sort(key=lambda x:str(x[0][x[1]] or '').strip())

    def set_aside_from_col(val):
        s = str(val or '').strip().replace('%','')
        if s == '30': return '30%'
        if s == '50': return '50%'
        if s == '60': return '60%'
        return str(val or '').strip() if val else ''

    def set_aside(unit_type):
        code=str(unit_type or '').strip()
        if not code: return ''
        m=re.search(r'(1|2)(50|60)',code)
        if m:
            pct=m.group(2)
            if pct=='50': return '50%'
            if pct=='60': return '60%'
        last=code[-1].upper()
        if last=='3': return '30%'
        if last=='5': return '50%'
        if last=='6': return '60%'
        if last=='M': return 'Exempt Unit'
        return ''

    def write_rr_row(rn, row, o, fc):
        unit=str(row[o] or '').strip()
        ut=str(row[o+1] or '').strip() if len(row)>o+1 else ''

        if is_vaf_format:
            sa_display = set_aside(ut)
            rname = str(row[o+4] or '').strip() if len(row)>o+4 else ''
            sq    = row[o+2]  if len(row)>o+2  else None
            mr    = row[o+5]  if len(row)>o+5  else None
            lg    = None
            sr    = None
            tr    = row[o+6]  if len(row)>o+6  else None
            lr    = row[o+6]  if len(row)>o+6  else None
            vac_r = None
            dep   = row[o+7]  if len(row)>o+7  else None
            mi    = row[o+9]  if len(row)>o+9  else None
            lf    = row[o+9]  if len(row)>o+9  else None
            lt    = row[o+10] if len(row)>o+10 else None
        else:
            sa_display = set_aside(ut)
            rname = str(row[o+2] or '').strip() if len(row)>o+2 else ''
            sq    = row[o+3]  if len(row)>o+3  else None
            mr    = row[o+4]  if len(row)>o+4  else None
            lg    = row[o+5]  if len(row)>o+5  else None
            sr    = row[o+6]  if len(row)>o+6  else None
            tr    = row[o+7]  if len(row)>o+7  else None
            lr    = row[o+8]  if len(row)>o+8  else None
            vac_r = row[o+9]  if len(row)>o+9  else None
            dep   = row[o+10] if len(row)>o+10 else None
            mi    = row[o+11] if len(row)>o+11 else None
            lf    = row[o+12] if len(row)>o+12 else None
            lt    = row[o+13] if len(row)>o+13 else None

        isvac = rname.strip().upper() in ('VACANT',' VACANT') or not rname.strip()

        def sc(col,val,h='left',fmt=None):
            cell=ws.cell(rn,col); cell.value=val; cell.font=gfont(color=fc); cell.fill=gfill(WHITE)
            cell.alignment=Alignment(horizontal=h,vertical='center',wrap_text=False)
            if fmt: cell.number_format=fmt

        sc(1,unit)
        sc(2,ut)
        sc(3,sa_display,'center')
        sc(4,' VACANT' if isvac else rname)
        sc(5,sq or 0,'center','#,##0')
        sc(6,mr or 0,'right','#,##0.00')
        try: lg_val=0 if isvac else float(str(lg or 0).replace(',',''))
        except: lg_val=0
        sc(7,lg_val,'right','#,##0.00')
        try: sr_val=float(str(sr or 0).replace(',',''))
        except: sr_val=0
        sc(8,sr_val,'right','#,##0.00')
        try: tr_val=float(str(tr or 0).replace(',',''))
        except: tr_val=0
        sc(9,tr_val,'right','#,##0.00')
        try: lr_val=float(str(lr or 0).replace(',',''))
        except: lr_val=0
        sc(10,lr_val,'right','#,##0.00')
        try: vac_val=-(float(str(mr or 0).replace(',',''))) if isvac else 0
        except: vac_val=0
        sc(11,vac_val,'right','#,##0.00')
        try: dep_val=float(str(dep or 0).replace(',',''))
        except: dep_val=0
        sc(12,dep_val,'right','#,##0.00')
        for col,dv in [(13,mi),(14,lf),(15,lt)]:
            cell=ws.cell(rn,col); cell.value=dv; cell.font=gfont(color=fc); cell.fill=gfill(WHITE)
            cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=False)
            if dv is not None and (isinstance(dv,_dt.datetime) or (isinstance(dv,(int,float)) and dv>40000)):
                cell.number_format='MM/DD/YY'
        ws.row_dimensions[rn].height=15.0

    rn=6
    for row,o in V: write_rr_row(rn,row,o,RED_FONT); rn+=1
    for row,o in O: write_rr_row(rn,row,o,BLACK); rn+=1
    data_end=rn-1

    for c in range(1,MC+1):
        ws.cell(rn,c).fill=gfill(GRAY_HDR); ws.cell(rn,c).font=gfont(bold=True)
    for c in range(5,13):
        ws.cell(rn,c).value=f'=SUM({chr(64+c)}6:{chr(64+c)}{data_end})'
        ws.cell(rn,c).number_format='#,##0.00'
        ws.cell(rn,c).alignment=Alignment(horizontal='right',vertical='center')
    rn+=1

    rn+=1
    for c in range(1,MC+1): ws.cell(rn,c).fill=gfill(GREEN)
    ws.cell(rn,1).value='Non-Revenue Units'; ws.cell(rn,1).font=gfont(bold=True,color='FF000000')
    ws.merge_cells(start_row=rn,start_column=1,end_row=rn,end_column=MC)
    rn+=1
    ws.cell(rn,4).value='No Data Available'; ws.cell(rn,4).font=gfont()
    rn+=2

    occ_sq=occ_mr=occ_sr=occ_tr=occ_dep=0; occ_cnt=0
    vac_sq=vac_mr=0; vac_cnt=len(V)
    for row,o in O:
        sq_idx  = o+2 if is_vaf_format else o+3
        mr_idx  = o+5 if is_vaf_format else o+4
        sr_idx  = -1  if is_vaf_format else o+6
        tr_idx  = o+6 if is_vaf_format else o+7
        dep_idx = o+7 if is_vaf_format else o+10
        try: occ_sq+=float(row[sq_idx] or 0)
        except: pass
        try: occ_mr+=float(str(row[mr_idx] or 0).replace(',',''))
        except: pass
        try: occ_sr+=float(str(row[sr_idx] or 0).replace(',','')) if sr_idx >= 0 else 0
        except: pass
        try: occ_tr+=float(str(row[tr_idx] or 0).replace(',',''))
        except: pass
        try: occ_dep+=float(str(row[dep_idx] or 0).replace(',',''))
        except: pass
        occ_cnt+=1
    for row,o in V:
        sq_idx = o+2 if is_vaf_format else o+3
        mr_idx = o+5 if is_vaf_format else o+4
        try: vac_sq+=float(row[sq_idx] or 0)
        except: pass
        try: vac_mr+=float(str(row[mr_idx] or 0).replace(',',''))
        except: pass
    tot_sq=occ_sq+vac_sq; tot_mr=occ_mr+vac_mr; tot_cnt=occ_cnt+vac_cnt

    ws.cell(rn,1).value='Total Market Rent :'; ws.cell(rn,1).font=gfont(bold=True)
    ws.cell(rn,2).value=tot_mr; ws.cell(rn,2).font=gfont(); ws.cell(rn,2).number_format='#,##0.00'
    ws.cell(rn,7).value='Total Potential Rent :'; ws.cell(rn,7).font=gfont(bold=True)
    ws.cell(rn,9).value=occ_tr; ws.cell(rn,9).font=gfont(); ws.cell(rn,9).number_format='#,##0.00'
    rn+=2

    sum_hdrs=['Square\nFootage','Market\nRent','Sub rent','Actual\nRent','Security\nDeposit','Other\nDeposit','# of\nUnits','Occupancy']
    for i,h in enumerate(sum_hdrs,5):
        ws.cell(rn,i).value=h; ws.cell(rn,i).font=gfont(bold=True)
        ws.cell(rn,i).fill=gfill(GRAY_HDR)
        ws.cell(rn,i).alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws.row_dimensions[rn].height=28; rn+=1

    tot_pct=f'{occ_cnt/tot_cnt*100:.2f}%' if tot_cnt else '0%'
    vac_pct=f'{vac_cnt/tot_cnt*100:.2f}%' if tot_cnt else '0%'
    for label,sq,mr,sr,tr,dep,cnt,pct in [
        ('Occupied Units',occ_sq,occ_mr,occ_sr,occ_tr,occ_dep,occ_cnt,tot_pct),
        ('Vacant Units',  vac_sq,vac_mr,0,0,0,vac_cnt,vac_pct),
        ('Totals:',       tot_sq,tot_mr,occ_sr,occ_tr,occ_dep,tot_cnt,'100%'),
    ]:
        bold=(label=='Totals:')
        ws.cell(rn,4).value=label; ws.cell(rn,4).font=gfont(bold=bold)
        for ci,val in [(5,sq),(6,mr),(7,sr),(8,tr),(9,dep),(10,0)]:
            ws.cell(rn,ci).value=val or 0; ws.cell(rn,ci).font=gfont(bold=bold)
            ws.cell(rn,ci).number_format='#,##0.00'
            ws.cell(rn,ci).alignment=Alignment(horizontal='right',vertical='center')
        ws.cell(rn,11).value=cnt; ws.cell(rn,11).font=gfont(bold=bold)
        ws.cell(rn,12).value=pct; ws.cell(rn,12).font=gfont(bold=bold)
        ws.cell(rn,12).alignment=Alignment(horizontal='right',vertical='center')
        for ci in range(4,13): ws.cell(rn,ci).border=Border(top=T,bottom=T,left=T,right=T)
        rn+=1

    for col,w in {'A':9,'B':12,'C':13,'D':22,'E':7,'F':11,'G':11,'H':9,'I':10,'J':10,'K':10,'L':10,'M':10,'N':10,'O':10}.items():
        ws.column_dimensions[col].width=w
    for r in range(1,5): ws.row_dimensions[r].height=14
    ws.freeze_panes='A6'
    return ws, len(V), len(O)

# ============================================================================
# EXPIRING LEASES (120 days) - new in v9.49
# ============================================================================
def parse_expiring(raw_bytes):
    """Parse Yardi Expiring Leases export. Returns list of dict rows, sorted oldest -> newest.
    Robust to variations in header text, header position, and date format (text or datetime)."""
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    ws = wb.active
    headers, rows = None, []
    header_keywords = ('lease expires', 'lease expiration', 'expires', 'lease end', 'expiration date')
    
    def parse_date(val):
        """Convert val to datetime if possible. Handles datetime, date, and various text formats."""
        if val is None:
            return None
        if isinstance(val, datetime):
            return val
        # Handle date object (without time)
        try:
            from datetime import date as date_cls
            if isinstance(val, date_cls):
                return datetime(val.year, val.month, val.day)
        except: pass
        if isinstance(val, str):
            s = val.strip()
            if not s:
                return None
            # Try several common date formats
            for fmt in ('%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d', '%m-%d-%Y', '%m-%d-%y',
                        '%d/%m/%Y', '%b %d, %Y', '%B %d, %Y', '%m/%d/%Y %H:%M:%S'):
                try:
                    return datetime.strptime(s, fmt)
                except: pass
        return None
    
    for r in ws.iter_rows(values_only=True):
        if headers is None:
            # Look for any cell in the row that matches one of the header keywords
            if r:
                row_lower = [str(c).strip().lower() if c is not None else '' for c in r]
                if any(any(kw in cell for kw in header_keywords) for cell in row_lower):
                    headers = [str(h).strip() if h is not None else '' for h in r]
                    # Find the index of the "Lease Expires" / similar column
                    for i, cell in enumerate(row_lower):
                        if any(kw in cell for kw in header_keywords):
                            expires_col_idx = i
                            break
                    else:
                        expires_col_idx = 0
            continue
        if not r or all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
            continue
        # Skip total/summary rows: check the first few columns for signal words
        for chk_idx in range(min(3, len(r))):
            chk = r[chk_idx]
            if isinstance(chk, str) and any(k in chk.lower() for k in ('total', 'grand', 'summary', 'village at')):
                break
        else:
            # No skip-keyword found in first 3 cols; try to parse the date
            date_val = r[expires_col_idx] if expires_col_idx < len(r) else None
            parsed = parse_date(date_val)
            if parsed is None:
                continue
            row_dict = {headers[i]: r[i] if i < len(r) else None for i in range(len(headers))}
            # Normalize: ensure 'Lease Expires' key exists with parsed datetime
            row_dict['Lease Expires'] = parsed
            rows.append(row_dict)
    wb.close()
    rows.sort(key=lambda x: x.get('Lease Expires') or datetime.max)
    return rows


def build_monthly_counts(rows):
    """Returns list of (month_label, count) sorted oldest -> newest."""
    counts = OrderedDict()
    for r in rows:
        d = r.get('Lease Expires')
        if not isinstance(d, datetime):
            continue
        key = d.strftime('%b %Y')
        sk = (d.year, d.month)
        if key not in counts:
            counts[key] = {'sort': sk, 'count': 0}
        counts[key]['count'] += 1
    return [(k, v['count']) for k, v in sorted(counts.items(), key=lambda kv: kv[1]['sort'])]


def fmt_expiring(wb_out, raw_bytes, date, prop):
    """Build the formatted 'Expiring Leases (120 days) MM.DD.YY' tab."""
    rows = parse_expiring(raw_bytes)
    out_cols = ['Lease Expires','Unit','Resident','Market Rent','Current Rent',
                'Loss to Lease','Current Lease Term','Months At Property',
                'MTM?','Appr Status','Comments']
    n_cols = len(out_cols)
    last_letter = get_column_letter(n_cols)

    tab = f'Expiring Leases {date}'
    if tab in wb_out.sheetnames: del wb_out[tab]
    ws = wb_out.create_sheet(tab)

    # Title rows 1-3, green
    for r in range(1, 4):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
        for c in range(1, n_cols+1):
            cell = ws.cell(r, c)
            cell.fill = gfill(GREEN)
            cell.font = gfont(bold=(r != 3), color=DARKGRAY)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(1, 1).value = prop
    ws.cell(2, 1).value = 'Expiring Leases (120 days)'
    ws.cell(3, 1).value = f'As Of: {date}'

    # Header row 5 (gray), with row 4 as gray spacer band
    for c in range(1, n_cols+1):
        cell = ws.cell(4, c)
        cell.fill = gfill(GRAY_AR)
        cell.font = gfont(bold=True)
        cell.border = bblack()
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for i, label in enumerate(out_cols, start=1):
        cell = ws.cell(5, i, value=label)
        cell.fill = gfill(GRAY_AR)
        cell.font = gfont(bold=True)
        cell.border = bblack()
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows
    start_row = 6
    for i, row in enumerate(rows):
        r = start_row + i
        vals = [
            row.get('Lease Expires'),
            row.get('Unit'),
            row.get('Resident'),
            row.get('Market Rent'),
            row.get('Current Rent'),
            row.get('Loss to Lease'),
            row.get('Current Lease Term'),
            row.get('Months At Property'),
            row.get('MTM?'),
            row.get('Appr Status'),
            row.get('Comments ') or row.get('Comments') or None,
        ]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(r, ci, value=v)
            cell.font = gfont()
            cell.fill = gfill(WHITE)
            cell.border = bgray()
            label = out_cols[ci-1]
            if label == 'Lease Expires':
                cell.number_format = 'MM/DD/YY'
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif label in ('Market Rent','Current Rent','Loss to Lease'):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif label in ('Unit','Current Lease Term','Months At Property','MTM?','Appr Status'):
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[r].height = 15.0

    # Totals row
    last_data_row = start_row + len(rows) - 1
    total_row = last_data_row + 1 if rows else start_row

    for col in range(1, n_cols+1):
        cell = ws.cell(total_row, col)
        cell.fill = gfill(GRAY_AR)
        cell.font = gfont(bold=True)
        cell.border = bblack()
    ws.cell(total_row, 1).value = 'Total'
    ws.cell(total_row, 1).alignment = Alignment(horizontal='left', vertical='center')

    if rows:
        ws.cell(total_row, 3).value = f'=COUNTA(C{start_row}:C{last_data_row})'
        ws.cell(total_row, 3).alignment = Alignment(horizontal='center', vertical='center')
        for ci, label in enumerate(out_cols, start=1):
            if label in ('Market Rent','Current Rent','Loss to Lease'):
                cl = get_column_letter(ci)
                cell = ws.cell(total_row, ci, value=f'=SUM({cl}{start_row}:{cl}{last_data_row})')
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.fill = gfill(GRAY_AR)
                cell.font = gfont(bold=True)
                cell.border = bblack()

    # Column widths
    widths = {'Lease Expires':13,'Unit':9,'Resident':28,'Market Rent':13,'Current Rent':13,
              'Loss to Lease':13,'Current Lease Term':13,'Months At Property':13,
              'MTM?':8,'Appr Status':13,'Comments':40}
    for i, label in enumerate(out_cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(label, 12)

    ws.freeze_panes = 'A6'
    return ws, rows
# ============================================================================

TRAFFIC_SOURCES=[
    'Apartment List','Apartment Ratings','Brochure/Flyer','Google','Craigslist',
    'Drive-by','Other','GeoTargeting','Facebook','Google My Business','Locator',
    'Organic Social','CoStar','Paid Search','Paid Social','Property Website',
    'Referral - Current Resident','Referral - Former Resident','Remarketing',
    'Rent.','RentPath','Yelp','Zillow','Zumper'
]

def fmt_traffic(wb_out, raw_bytes, date, prop):
    import csv, io as _io
    text = raw_bytes.decode('utf-8-sig')
    reader = csv.DictReader(_io.StringIO(text))
    src_data = {}
    for row in reader:
        src = str(row.get('Source','') or '').strip()
        if not src or src.lower() == 'total': continue
        def iv(k):
            try: return int(str(row.get(k,0) or 0).replace(',',''))
            except: return 0
        src_data[src] = {
            'plan': str(row.get('Plan','') or '').strip(),
            'leads': iv('Leads'), 'prospects': iv('Prospects'),
            'visits': iv('Visits'), 'leases': iv('Leases'),
            'applications': iv('Applications')
        }

    import datetime as _dt
    tab = f'Weekly Traffic {date}'
    if tab in wb_out.sheetnames: del wb_out[tab]
    ws = wb_out.create_sheet(tab)

    try:
        d = _dt.datetime.strptime(date, '%m.%d.%y')
        mon = d - _dt.timedelta(days=d.weekday())
        sun = mon + _dt.timedelta(days=6)
        date_range = f"{mon.month}/{mon.day}/{mon.year}-{sun.month}/{sun.day}/{sun.year}"
    except:
        date_range = date

    MC = 7
    titles = [
        ('Weekly Traffic', True),
        (prop, False),
        (date_range, True),
    ]
    for ti, (txt, bold) in enumerate(titles):
        r = ti + 1
        for c in range(1, MC+1):
            ws.cell(r,c).fill = gfill(GREEN)
            ws.cell(r,c).font = gfont(bold=bold)
            ws.cell(r,c).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(r,1).value = txt
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MC)
        ws.row_dimensions[r].height = 15.0

    hdrs = ['Source','Plan','Leads','Prospects','Visits','Leases','Applications']
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(4, c)
        cell.value = h; cell.font = gfont(bold=True); cell.fill = gfill(GRAY_HDR)
        cell.alignment = Alignment(horizontal='center' if c > 1 else 'left', vertical='center')
        cell.border = bblack()
    ws.row_dimensions[4].height = 15.0

    rn = 5
    for src in TRAFFIC_SOURCES:
        d = src_data.get(src, {'plan':'','leads':0,'prospects':0,'visits':0,'leases':0,'applications':0})
        ws.cell(rn,1).value = src
        ws.cell(rn,2).value = d['plan'] or None
        for ci, key in enumerate(['leads','prospects','visits','leases','applications'], 3):
            ws.cell(rn,ci).value = d[key]
            ws.cell(rn,ci).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(rn,ci).number_format = '0'
        for c in range(1, MC+1):
            ws.cell(rn,c).font = gfont()
            ws.cell(rn,c).fill = gfill(WHITE)
            ws.cell(rn,c).border = Border(top=TG, bottom=TG, left=TG, right=TG)
        ws.cell(rn,1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(rn,2).alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[rn].height = 15.0
        rn += 1

    total_row = rn
    ws.cell(rn,1).value = 'Total'
    ws.cell(rn,1).font = gfont(bold=True)
    ws.cell(rn,1).fill = gfill(GRAY_HDR)
    ws.cell(rn,1).border = bblack()
    ws.cell(rn,2).fill = gfill(GRAY_HDR); ws.cell(rn,2).border = bblack()
    data_start = 5
    for c in range(3, MC+1):
        cell = ws.cell(rn, c)
        cell.value = f'=SUM({get_column_letter(c)}{data_start}:{get_column_letter(c)}{rn-1})'
        cell.font = gfont(bold=True); cell.fill = gfill(GRAY_HDR)
        cell.border = bblack()
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.number_format = '0'
    ws.row_dimensions[rn].height = 15.0

    for col, w in {'A':28,'B':10,'C':10,'D':12,'E':10,'F':10,'G':14}.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'A5'
    return ws

def parse_traffic(raw_bytes, date):
    import csv, io as _io
    rows = []
    try:
        text = raw_bytes.decode('utf-8-sig')
        reader = csv.reader(_io.StringIO(text))
        rows = list(reader)
    except Exception:
        try:
            wb_r = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
            for r in wb_r.active.iter_rows(values_only=True):
                rows.append([str(c or '') for c in r])
            wb_r.close()
        except Exception:
            return None

    start_date = end_date = ''
    for row in rows[:10]:
        for i, cell in enumerate(row):
            if 'startdate' in str(cell).lower().replace(' ','') and i+1 < len(row):
                start_date = str(row[i+1]).strip()
            if 'enddate' in str(cell).lower().replace(' ','') and i+1 < len(row):
                end_date = str(row[i+1]).strip()
    def fmt_d(d):
        try:
            from datetime import datetime
            return datetime.strptime(d, '%Y-%m-%d').strftime('%-m/%-d/%Y')
        except: return d
    date_range = f'{fmt_d(start_date)} - {fmt_d(end_date)}' if start_date and end_date else date

    hi = next((i for i,r in enumerate(rows) if r and 'source' in str(r[0]).lower() and any('lead' in str(c).lower() for c in r)), 0)
    hdr = rows[hi] if rows else []
    def ci(name):
        for i,h in enumerate(hdr):
            if name.lower() in str(h).lower(): return i
        return -1
    c_leads=ci('lead'); c_pros=ci('prospect'); c_vis=ci('visit'); c_lease=ci('lease'); c_app=ci('applic')

    SOURCES = [
        'Apartment List','Apartment Ratings','Brochure/Flyer','Google','Craigslist',
        'Drive-by','Other','GeoTargeting','Facebook','Google My Business','Locator',
        'Organic Social','CoStar','Paid Search','Paid Social','Property Website',
        'Referral - Current Resident','Referral - Former Resident','Remarketing',
        'Rent.','RentPath','Yelp','Zillow','Zumper'
    ]
    data = {}
    for row in rows[hi+1:]:
        if not row or not str(row[0]).strip(): continue
        src = str(row[0]).strip()
        if src not in SOURCES: continue
        def gv(idx):
            if idx < 0 or idx >= len(row): return 0
            v = str(row[idx]).strip().replace(',','')
            try: return int(float(v))
            except: return 0
        data[src] = [gv(c_leads), gv(c_pros), gv(c_vis), gv(c_lease), gv(c_app)]

    active = [(src, data[src]) for src in SOURCES if src in data and any(v!=0 for v in data[src])]
    return {'date_range': date_range, 'rows': active}

# ============================================================================
# REPUTATION (Opinionn Review Summary PDF) - new in v9.50
# ============================================================================
def parse_review_summary_pdf(raw_bytes):
    """Parse the Opinionn 'Review Summary' PDF. Returns dict with overall and per-platform metrics.
    
    Expected PDF structure (from Reports -> Review Summary):
      - Top: 'Review Summary' / 'Village at Madrone - FPI Management' / date range / 'All Sources'
      - Cards row: 'Overall Public Rating' / '3.4' / '70 Reviews', then per-platform cards
        (Google, Yelp, Apartments.com, Facebook, ApartmentRatings.com, Opiniion)
      - 'Reviews by Month' chart
      - Source table: Source | Average Rating | All Time Rating | Positive Reviews | Negative Reviews | Total Reviews
    
    Returns:
      {
        'overall_rating': float,
        'overall_total': int,
        'platforms': {
          'Google': {'avg_rating': float, 'all_time_rating': float, 'positive': int, 'negative': int, 'total': int},
          ...
        }
      }
    
    Returns None if parsing fails for any reason. The block builder handles None gracefully.
    """
    try:
        import pdfplumber
    except ImportError:
        return None
    
    PLATFORMS = ['Google', 'Apartments.com', 'Yelp', 'Facebook', 'ApartmentRatings.com', 'Opiniion']
    
    result = {
        'overall_rating': 0.0,
        'overall_total': 0,
        'platforms': {p: {'avg_rating': 0.0, 'all_time_rating': 0.0,
                          'positive': 0, 'negative': 0, 'total': 0} for p in PLATFORMS}
    }
    
    try:
        with pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
            if not pdf.pages:
                return None
            page = pdf.pages[0]
            text = page.extract_text() or ''
            
            # ===== Parse the cards section =====
            # Layout: card title on one line, then number with optional star, then "X Reviews"
            # Examples from real PDF:
            #   Overall Public Rating
            #   3.4
            #   70 Reviews
            #   Google
            #   3.5
            #   53 Reviews
            
            lines = [ln.strip() for ln in text.split('\n') if ln.strip()]
            
            # Walk through lines looking for card patterns
            i = 0
            while i < len(lines):
                line = lines[i]
                # Overall Public Rating card
                if re.search(r'overall\s*public\s*rating', line, re.I):
                    # Next non-empty line should have the rating, then 'X Reviews'
                    rating, total = _extract_card_values(lines, i + 1)
                    if rating is not None:
                        result['overall_rating'] = rating
                    if total is not None:
                        result['overall_total'] = total
                    i += 1
                    continue
                # Per-platform cards
                matched = False
                for plat in PLATFORMS:
                    # Match exact platform name (case-insensitive). Use word boundaries
                    # to avoid 'Apartments.com' matching inside 'ApartmentRatings.com'.
                    if line.lower() == plat.lower() or re.fullmatch(rf'\s*{re.escape(plat)}\s*', line, re.I):
                        rating, total = _extract_card_values(lines, i + 1)
                        if rating is not None:
                            result['platforms'][plat]['avg_rating'] = rating
                        if total is not None:
                            result['platforms'][plat]['total'] = total
                        matched = True
                        i += 1
                        break
                if matched:
                    continue
                i += 1
            
            # ===== Parse the source table =====
            # Try extracting tables. The Source breakdown table typically has columns:
            # Source | Average Rating | All Time Rating | Positive Reviews | Negative Reviews | Total Reviews
            try:
                tables = page.extract_tables()
                for tbl in tables or []:
                    if not tbl or len(tbl) < 2:
                        continue
                    # Find a header row that mentions 'Source' and 'Total Reviews' or similar
                    header_row_idx = None
                    for hi, row in enumerate(tbl):
                        if row and any(c and 'source' in str(c).lower() for c in row):
                            if any(c and 'total' in str(c).lower() for c in row):
                                header_row_idx = hi
                                break
                    if header_row_idx is None:
                        continue
                    
                    # Map column indices
                    header = [str(c or '').strip().lower() for c in tbl[header_row_idx]]
                    col_idx = {}
                    for ci, h in enumerate(header):
                        if 'source' in h: col_idx['source'] = ci
                        elif 'average rating' in h or h == 'avg rating': col_idx['avg'] = ci
                        elif 'all time' in h: col_idx['all_time'] = ci
                        elif 'positive' in h: col_idx['positive'] = ci
                        elif 'negative' in h: col_idx['negative'] = ci
                        elif 'total' in h: col_idx['total'] = ci
                    
                    # Walk the data rows
                    for row in tbl[header_row_idx + 1:]:
                        if not row or 'source' not in col_idx:
                            continue
                        src_cell = row[col_idx['source']] if col_idx['source'] < len(row) else None
                        if not src_cell:
                            continue
                        src_name = str(src_cell).strip()
                        # Match against canonical platform names
                        matched_plat = None
                        for plat in PLATFORMS:
                            if plat.lower() in src_name.lower() or src_name.lower() in plat.lower():
                                matched_plat = plat
                                break
                        if not matched_plat:
                            continue
                        # Pull numeric values, treating dashes/blanks as 0
                        if 'avg' in col_idx and col_idx['avg'] < len(row):
                            v = _parse_num(row[col_idx['avg']])
                            if v is not None:
                                result['platforms'][matched_plat]['avg_rating'] = v
                        if 'all_time' in col_idx and col_idx['all_time'] < len(row):
                            v = _parse_num(row[col_idx['all_time']])
                            if v is not None:
                                result['platforms'][matched_plat]['all_time_rating'] = v
                        if 'positive' in col_idx and col_idx['positive'] < len(row):
                            v = _parse_num(row[col_idx['positive']])
                            if v is not None:
                                result['platforms'][matched_plat]['positive'] = int(v)
                        if 'negative' in col_idx and col_idx['negative'] < len(row):
                            v = _parse_num(row[col_idx['negative']])
                            if v is not None:
                                result['platforms'][matched_plat]['negative'] = int(v)
                        if 'total' in col_idx and col_idx['total'] < len(row):
                            v = _parse_num(row[col_idx['total']])
                            if v is not None and v > 0:
                                result['platforms'][matched_plat]['total'] = int(v)
                    break  # only process the first matching source table
            except Exception:
                pass  # table extraction is best-effort; cards are primary source
        
        return result
    except Exception:
        traceback.print_exc()
        return None


def _extract_card_values(lines, start_idx, max_lookahead=4):
    """Helper: starting at lines[start_idx], find (rating, total) values from a card.
    Looks ahead up to max_lookahead lines for a number (rating) and 'X Reviews' (total).
    Returns (rating_or_None, total_or_None)."""
    rating = None
    total = None
    end = min(start_idx + max_lookahead, len(lines))
    for j in range(start_idx, end):
        ln = lines[j].strip()
        if not ln:
            continue
        # Look for "X Reviews" pattern (allow optional decimal but expect integer)
        m_rev = re.match(r'^([\d,]+)\s*reviews?\b', ln, re.I)
        if m_rev and total is None:
            try:
                total = int(m_rev.group(1).replace(',', ''))
            except: pass
            continue
        # Look for a standalone rating number (1-5 range, may have decimals, may have trailing star/info icon char)
        m_rate = re.match(r'^([0-5](?:\.[0-9]+)?)\s*[\u2605\*\u24d8]?\s*$', ln)
        if m_rate and rating is None:
            try:
                rating = float(m_rate.group(1))
            except: pass
            continue
        # If the line has "Industry Average" or similar metadata, skip but don't break
        if 'industry' in ln.lower() or 'average' in ln.lower():
            continue
        # If we've found both, stop
        if rating is not None and total is not None:
            break
    return rating, total


def _parse_num(val):
    """Parse a numeric value from a table cell, treating dashes/blanks as None.
    Returns float or None."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s == '-' or s.lower() in ('n/a', 'na'):
        return None
    # Strip commas, percentage signs, currency signs
    s = s.replace(',', '').replace('$', '').replace('%', '').strip()
    try:
        return float(s)
    except:
        return None


def build_reputation_block(ws, reputation_data, start_row=34, start_col=5):
    """Build the Reputation block at E34:K43 (default) on the Weekly Summary tab.
    
    Layout:
      Row 34: 'Reputation' title bar (merged E:K)
      Row 35: Overall summary, label/value/label/value/label/value/blank (E F G H I J K)
              E='Overall Public Rating', F=rating value, G='Total Reviews', H=count, I='Industry Avg', J=3.9, K=blank
      Row 36: 'Per Platform Breakdown' sub-header (merged E:K)
      Row 37: Column headers (Source, Average Rating, All Time Rating, Positive Reviews, Negative Reviews, Total Reviews, Notes)
      Rows 38-43: 6 platforms (Google, Apartments.com, Yelp, Facebook, ApartmentRatings.com, Opiniion)
    
    If reputation_data is None, blank input cells are rendered for manual entry.
    """
    BLUE = 'FFBDD7EE'  # match the unified blue palette used elsewhere on Weekly Summary
    AB = bblack()
    f9 = gfont(sz=9)
    f9b = gfont(bold=True, sz=9)
    
    # Default empty data shape
    if reputation_data is None:
        reputation_data = {
            'overall_rating': None,
            'overall_total': None,
            'platforms': {}
        }
    
    PLATFORMS = ['Google', 'Apartments.com', 'Yelp', 'Facebook', 'ApartmentRatings.com', 'Opiniion']
    
    # Block spans 7 columns: E, F, G, H, I, J, K (cols 5-11)
    end_col = start_col + 6  # K = 11
    
    def style_cell(r, c, val=None, fill=None, bold=False, align='left', fmt=None):
        cell = ws.cell(r, c)
        if val is not None:
            cell.value = val
        if fill:
            cell.fill = gfill(fill)
        cell.font = gfont(bold=bold, sz=9)
        cell.alignment = Alignment(horizontal=align, vertical='center')
        if fmt:
            cell.number_format = fmt
        cell.border = AB
        return cell
    
    # ===== Row 34: Title bar (merged E:K) =====
    ws.merge_cells(start_row=start_row, start_column=start_col,
                   end_row=start_row, end_column=end_col)
    for c in range(start_col, end_col + 1):
        cell = ws.cell(start_row, c)
        cell.fill = gfill(BLUE)
        cell.font = f9b
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = AB
    ws.cell(start_row, start_col).value = 'Reputation'
    
    # ===== Row 35: Overall summary =====
    r = start_row + 1
    overall_layout = [
        (start_col,     'Overall Public Rating', 'label'),                                        # E
        (start_col + 1, reputation_data.get('overall_rating'), 'rating'),                          # F
        (start_col + 2, 'Total Reviews', 'label'),                                                # G
        (start_col + 3, reputation_data.get('overall_total'), 'count'),                            # H
        (start_col + 4, 'Industry Avg', 'label'),                                                 # I
        (start_col + 5, 3.9, 'rating'),                                                           # J
    ]
    for col, val, kind in overall_layout:
        cell = ws.cell(r, col)
        cell.border = AB
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if kind == 'label':
            cell.value = val
            cell.font = f9b
            cell.fill = gfill(WHITE)
        else:
            # input value (blue fill)
            cell.value = val if val is not None else 0
            cell.font = f9b
            cell.fill = gfill(BLUE)
            cell.number_format = '0.0' if kind == 'rating' else '0'
    # Column K (end_col) on row 35 stays blank but bordered
    ws.cell(r, end_col).border = AB
    ws.cell(r, end_col).fill = gfill(WHITE)
    
    # ===== Row 36: 'Per Platform Breakdown' sub-header (merged E:K) =====
    r = start_row + 2
    ws.merge_cells(start_row=r, start_column=start_col,
                   end_row=r, end_column=end_col)
    for c in range(start_col, end_col + 1):
        cell = ws.cell(r, c)
        cell.fill = gfill(BLUE)
        cell.font = f9b
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = AB
    ws.cell(r, start_col).value = 'Per Platform Breakdown'
    
    # ===== Row 37: Column headers =====
    r = start_row + 3
    headers = ['Source', 'Average Rating', 'All Time Rating',
               'Positive Reviews', 'Negative Reviews', 'Total Reviews', 'Notes']
    for i, h in enumerate(headers):
        style_cell(r, start_col + i, val=h, fill=BLUE, bold=True, align='center')
    
    # ===== Rows 38-43: Platform data rows =====
    for i, plat in enumerate(PLATFORMS):
        r = start_row + 4 + i
        plat_data = reputation_data['platforms'].get(plat, {})
        
        # Source name (bold, left-aligned)
        src_cell = ws.cell(r, start_col)
        src_cell.value = plat
        src_cell.font = f9b
        src_cell.alignment = Alignment(horizontal='left', vertical='center')
        src_cell.border = AB
        src_cell.fill = gfill(WHITE)
        
        # Numeric cells (blue input fill)
        numeric_cols = [
            (start_col + 1, plat_data.get('avg_rating', 0), 'rating'),
            (start_col + 2, plat_data.get('all_time_rating', 0), 'rating'),
            (start_col + 3, plat_data.get('positive', 0), 'count'),
            (start_col + 4, plat_data.get('negative', 0), 'count'),
            (start_col + 5, plat_data.get('total', 0), 'count'),
        ]
        for col, val, kind in numeric_cols:
            cell = ws.cell(r, col)
            cell.value = val if val is not None else 0
            cell.font = f9
            cell.fill = gfill(BLUE)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = AB
            cell.number_format = '0.0' if kind == 'rating' else '0'
        
        # Notes column (free-text, white)
        notes_cell = ws.cell(r, end_col)
        notes_cell.value = None
        notes_cell.font = f9
        notes_cell.alignment = Alignment(horizontal='left', vertical='center')
        notes_cell.border = AB
        notes_cell.fill = gfill(WHITE)
# ============================================================================


def build_weekly_summary(wb_out, wb_ro, date, prop, ua_ws=None, tar_ws=None, sar_ws=None, tar_total=0, sar_total=0, rr_ws=None, traffic_data=None, expiring_rows=None, tar_red_count=0, reputation_data=None):
    """
    Build Weekly Summary per spec:
      - 10 cols wide (A-J), ~56 rows
      - Calibri 9pt, fill default white, accent #BDD7EE on headers/totals/inputs
      - Col A = left-margin spacer with right border (A1:A56)
      - Sections 1-8 (see spec)
      - Section 8 'gutter trick': D25-56 cleared with specific borders kept
    """
    BLUE = 'FFBDD7EE'
    
    total_units = get_total_units(prop)
    
    # Read NTV data from source workbook
    ws_name = next((n for n in wb_ro.sheetnames if 'weekly summary' in n.lower()), None)
    src_vals = {}
    if ws_name:
        ws_src = wb_ro[ws_name]
        for row in ws_src.iter_rows(values_only=False):
            for cell in row:
                if cell.value is not None:
                    src_vals[(cell.row, cell.column)] = cell.value
        if ws_name in wb_out.sheetnames:
            del wb_out[ws_name]
        ws = wb_out.create_sheet(ws_name)
    else:
        ws = wb_out.create_sheet('Weekly Summary ')
    
    f9 = gfont(sz=9)
    f9b = gfont(bold=True, sz=9)
    AB = bblack()  # all-sides thin black
    
    BR = Border(right=T)   # right only
    BL = Border(left=T)    # left only
    BLR = Border(left=T, right=T)  # left + right
    
    def style(r, c, val=None, fill=None, bold=False, border=None,
              align='left', fmt=None, wrap=False):
        cell = ws.cell(r, c)
        if val is not None: cell.value = val
        if fill: cell.fill = gfill(fill)
        cell.font = gfont(bold=bold, sz=9)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
        if fmt: cell.number_format = fmt
        if border: cell.border = border
        return cell
    
    # =========================================================================
    # COL A — left-margin spacer with right border (A1:A56)
    # =========================================================================
    for r in range(1, 57):
        ws.cell(r, 1).border = BR
        ws.cell(r, 1).font = f9
    
    # =========================================================================
    # SECTION 1: Title block (rows 1-3) — merge B:H, blue, bold, centered, AB
    # =========================================================================
    title_lines = [
        src_vals.get((1, 2), prop.split('(')[0].strip()),
        src_vals.get((2, 2), 'Occupancy & Delinquency Summary'),
        date,
    ]
    for i, txt in enumerate(title_lines):
        r = i + 1
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        for c in range(2, 9):
            cell = ws.cell(r, c)
            cell.fill = gfill(BLUE)
            cell.font = f9b
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = AB
        ws.cell(r, 2).value = txt
    
    # =========================================================================
    # SECTION 2: Occupancy waterfall (rows 4-12)
    # =========================================================================
    # All cells in rows 4-12 across B:H get borders
    for r in range(4, 13):
        for c in range(2, 9):
            ws.cell(r, c).border = AB
            ws.cell(r, c).font = f9
    
    # Row 4: empty spacer with merged D:H
    ws.merge_cells(start_row=4, start_column=4, end_row=4, end_column=8)
    
    # Row 5: Total Units (B and C plain — NOT blue, NOT bold; D:H merged label)
    cell_b5 = ws.cell(5, 2); cell_b5.value = total_units
    cell_b5.font = f9
    cell_b5.alignment = galign('center'); cell_b5.border = AB
    cell_c5 = ws.cell(5, 3); cell_c5.value = '=B5/$B$5'
    cell_c5.font = f9
    cell_c5.alignment = galign('center'); cell_c5.number_format = '0.00%'; cell_c5.border = AB
    ws.merge_cells(start_row=5, start_column=4, end_row=5, end_column=8)
    style(5, 4, 'Total Units', align='left', border=AB)
    
    # Rows 6-11: Add/Subtract waterfall lines
    occ = [
        (6, 'Subtract', 'Physically Vacant'),
        (7, 'Add', 'Applications - Approved @ KG'),
        (8, 'Add', 'Applications - Pending Not Approved @ KG'),
        (9, 'Add', 'Applications - Site Processing - Not Sent to KG'),
        (10, 'Subtract', 'Notices to Vacate Not at Legal'),
        (11, 'Subtract', 'Notices to Vacate @ Legal'),
    ]
    for r, lbl, desc in occ:
        # Col A: Subtract/Add label (left of border, no fill)
        cell_a = ws.cell(r, 1); cell_a.value = lbl
        cell_a.font = f9; cell_a.alignment = galign('center')
        cell_a.border = BR  # just right border for column A
        # Col B: input cell (blue fill, no value yet, populated below from ua_ws)
        cell_b = ws.cell(r, 2)
        cell_b.fill = gfill(BLUE); cell_b.font = f9
        cell_b.alignment = galign('center'); cell_b.border = AB
        # Col C: percentage formula
        style(r, 3, f'=B{r}/$B$5', fmt='0.00%', align='center', border=AB)
        # Col D:H merged = description
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
        style(r, 4, desc, align='left', border=AB)
    
    # Row 12: NET LEASED (B and C plain — NOT blue, NOT bold; D:H merged label)
    cell_b12 = ws.cell(12, 2); cell_b12.value = '=B5+B6+B7+B8+B9+B10+B11'
    cell_b12.font = f9
    cell_b12.alignment = galign('center'); cell_b12.border = AB
    cell_c12 = ws.cell(12, 3); cell_c12.value = '=B12/$B$5'
    cell_c12.font = f9
    cell_c12.alignment = galign('center'); cell_c12.number_format = '0.00%'; cell_c12.border = AB
    ws.merge_cells(start_row=12, start_column=4, end_row=12, end_column=8)
    style(12, 4, 'NET LEASED ', align='left', border=AB)
    
    # =========================================================================
    # SECTION 3: Delinquency (rows 13-20) — mixed merge widths
    # =========================================================================
    # Row 13: blank spacer with merged D:H
    for c in range(2, 9):
        ws.cell(13, c).border = AB
        ws.cell(13, c).font = f9
    ws.merge_cells(start_row=13, start_column=4, end_row=13, end_column=8)
    
    # Row 14: # tenants owing | D:E label | F = eviction count | G = "@ legal" text
    for c in range(2, 9):
        ws.cell(14, c).border = AB
        ws.cell(14, c).font = f9
    cell_b14 = ws.cell(14, 2)
    cell_b14.fill = gfill(BLUE); cell_b14.font = f9
    cell_b14.alignment = galign('center'); cell_b14.border = AB
    style(14, 3, '=B14/B5', fmt='0.00%', align='center', border=AB)
    ws.merge_cells(start_row=14, start_column=4, end_row=14, end_column=5)
    style(14, 4, "# of tenants who haven't paid full rent", align='left', border=AB)
    style(14, 7, '@ legal', align='left', border=AB)
    
    # Row 15: blank spacer with merged D:H
    for c in range(2, 9):
        ws.cell(15, c).border = AB
        ws.cell(15, c).font = f9
    ws.merge_cells(start_row=15, start_column=4, end_row=15, end_column=8)
    
    # Row 16: # Physically Occupied | D:H merged label
    for c in range(2, 9):
        ws.cell(16, c).border = AB
        ws.cell(16, c).font = f9
    cell_b16 = ws.cell(16, 2)
    cell_b16.font = f9; cell_b16.alignment = galign('center'); cell_b16.border = AB
    cell_c16 = ws.cell(16, 3)
    cell_c16.fill = gfill(BLUE); cell_c16.font = f9
    cell_c16.alignment = galign('center')
    cell_c16.number_format = '#,##0_);(#,##0)'; cell_c16.border = AB
    ws.merge_cells(start_row=16, start_column=4, end_row=16, end_column=8)
    style(16, 4, '# Physically Occupied and Total Leased Rent', align='left', border=AB)
    
    # Row 17: blank spacer with merged D:H
    for c in range(2, 9):
        ws.cell(17, c).border = AB
        ws.cell(17, c).font = f9
    ws.merge_cells(start_row=17, start_column=4, end_row=17, end_column=8)
    
    # Rows 18-20: AR section (D:E merged label, F:H merged percentage)
    AR_FMT = '_([$$-409]* #,##0.00_);_([$$-409]* \\(#,##0.00\\);_([$$-409]* "-"??_);_(@_)'
    ar_rows = [
        (18, 'Tenant Accounts Receivable (AR)', '=C18/C16', True),
        (19, 'Subsidy Accounts Receivable (AR) ', '=C19/C16', True),
        (20, 'Total  AR', '=SUM(F18:F19)', False),
    ]
    for r, desc, pct_formula, blue_input in ar_rows:
        for c in range(2, 9):
            ws.cell(r, c).border = AB
            ws.cell(r, c).font = f9
        # Col B: empty
        ws.cell(r, 2).border = AB
        # Col C: AR currency value
        cell_c = ws.cell(r, 3)
        if blue_input:
            cell_c.fill = gfill(BLUE)
        cell_c.font = f9
        cell_c.alignment = galign('center')
        cell_c.number_format = AR_FMT
        cell_c.border = AB
        # D:E merged label
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        style(r, 4, desc, align='left', border=AB)
        # F:H merged percentage
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
        style(r, 6, pct_formula, fmt='0.00%', align='right', border=AB)
    # Total AR sum (C20)
    ws.cell(20, 3).value = '=SUM(C18:C19)'
    
    # =========================================================================
    # SECTION 4: Footnote (rows 21-22) — merge B:H both rows
    # =========================================================================
    ws.merge_cells(start_row=21, start_column=2, end_row=21, end_column=8)
    for c in range(2, 9):
        ws.cell(21, c).border = AB
        ws.cell(21, c).font = f9
    
    ws.merge_cells(start_row=22, start_column=2, end_row=22, end_column=8)
    for c in range(2, 9):
        ws.cell(22, c).border = AB
        ws.cell(22, c).font = f9
    style(22, 2, '* AR to include current month delinquency beginning 10th of each month', align='left', border=AB)
    
    # =========================================================================
    # SECTION 5: NTV table (rows 25-28) — B25:D28 box
    # =========================================================================
    # Row 25: NTV header merged B:C, blue+bold
    ws.merge_cells('B25:C25')
    style(25, 2, 'NTV', fill=BLUE, bold=True, align='center', border=AB)
    ws.cell(25, 3).fill = gfill(BLUE); ws.cell(25, 3).border = AB
    
    # Row 26: sub-headers
    style(26, 2, 'Unit #', fill=BLUE, bold=True, align='center', border=AB)
    style(26, 3, 'Move-in year', fill=BLUE, bold=True, align='center', border=AB)
    
    # Rows 27-28: NTV data (only border filled rows)
    for r in (27, 28):
        v2 = src_vals.get((r, 2)); v3 = src_vals.get((r, 3))
        if v2 or v3:
            cell_b = ws.cell(r, 2)
            if v2:
                cell_b.value = v2
                cell_b.alignment = galign('center')
            cell_b.font = f9; cell_b.border = AB
            cell_c = ws.cell(r, 3)
            if v3:
                cell_c.value = v3
                cell_c.number_format = 'MM/DD/YY'
                cell_c.alignment = galign('center')
            cell_c.font = f9; cell_c.border = AB
    
    # =========================================================================
    # SECTION 6: Weekly Traffic (rows 25-31, cols E-J)
    # =========================================================================
    if traffic_data:
        TR_COL = 5  # column E
        TR_COLS = 6
        tr_hdrs = ['Source', 'Leads', 'Prospects', 'Visits', 'Leases', 'Applications']
        active_sources = [(s, v) for s, v in traffic_data['rows'] if any(x != 0 for x in v)]
        date_range = traffic_data.get('date_range', date)
        
        # Row 25: Title strip merged E:J
        ws.merge_cells(start_row=25, start_column=TR_COL,
                       end_row=25, end_column=TR_COL + TR_COLS - 1)
        for c in range(TR_COL, TR_COL + TR_COLS):
            cell = ws.cell(25, c)
            cell.fill = gfill(BLUE); cell.font = f9b
            cell.alignment = galign('center'); cell.border = AB
        ws.cell(25, TR_COL).value = f'Weekly Traffic  " {date_range} "'
        
        # Row 26: sub-headers
        for i, h in enumerate(tr_hdrs):
            style(26, TR_COL + i, h, fill=BLUE, bold=True, align='center', border=AB)
        
        # Data rows
        rn = 27
        for src, vals in active_sources:
            style(rn, TR_COL, src, align='left', border=AB)
            for i, v in enumerate(vals):
                style(rn, TR_COL + 1 + i, v, align='center', border=AB, fmt='0')
            rn += 1
        
        # Total row
        tr_data_start = 27
        tr_data_end = rn - 1
        style(rn, TR_COL, 'Total', fill=BLUE, bold=True, align='center', border=AB)
        for i in range(1, TR_COLS):
            c = TR_COL + i
            cl = get_column_letter(c)
            style(rn, c, f'=SUM({cl}{tr_data_start}:{cl}{tr_data_end})',
                  fill=BLUE, bold=True, align='center', border=AB, fmt='0')
    
    # =========================================================================
    # SECTION 7: Expiring Leases (rows 33-56)
    # =========================================================================
    if expiring_rows:
        monthly = build_monthly_counts(expiring_rows)
        EX_COL = 2  # column B
        ex_start = 33
        
        # B33:C33 merged header
        ws.merge_cells(start_row=ex_start, start_column=EX_COL,
                       end_row=ex_start, end_column=EX_COL + 1)
        style(ex_start, EX_COL, 'Expiring Leases (120 days)',
              fill=BLUE, bold=True, align='center', border=AB)
        ws.cell(ex_start, EX_COL + 1).fill = gfill(BLUE)
        ws.cell(ex_start, EX_COL + 1).border = AB
        
        # Sub-headers
        sub_r = ex_start + 1
        style(sub_r, EX_COL, 'Month', fill=BLUE, bold=True, align='center', border=AB)
        style(sub_r, EX_COL + 1, 'Count', fill=BLUE, bold=True, align='center', border=AB)
        
        # Data
        data_start = sub_r + 1
        for i, (m, cnt) in enumerate(monthly):
            r = data_start + i
            style(r, EX_COL, m, align='center', border=AB)
            style(r, EX_COL + 1, cnt, align='center', border=AB, fmt='0')
        
        # Total
        total_r = data_start + len(monthly)
        last_data_r = total_r - 1 if monthly else data_start
        style(total_r, EX_COL, 'Total', fill=BLUE, bold=True, align='center', border=AB)
        if monthly:
            count_letter = get_column_letter(EX_COL + 1)
            style(total_r, EX_COL + 1,
                  f'=SUM({count_letter}{data_start}:{count_letter}{last_data_r})',
                  fill=BLUE, bold=True, align='center', border=AB, fmt='0')
        else:
            style(total_r, EX_COL + 1, 0, fill=BLUE, bold=True, align='center', border=AB, fmt='0')
    
    # =========================================================================
    # SECTION 8: Reputation block (rows 34-43, cols E-K) — new in v9.50
    # =========================================================================
    # Builds at E34:K43, in the same column band as Weekly Traffic so it sits
    # naturally below the traffic table. Pulls from Opinionn Review Summary PDF
    # via parse_review_summary_pdf(). If no PDF is uploaded or parsing fails,
    # reputation_data is None and the block renders with zeros for manual entry.
    build_reputation_block(ws, reputation_data, start_row=34, start_col=5)
    
    # =========================================================================
    # =========================================================================
    # (No gutter trick — column E is wide and holds Weekly Traffic data;
    #  no borders needed on column D or E in rows 25-56)
    # =========================================================================
    
    # =========================================================================
    # Compute and fill values from input data
    # =========================================================================
    ws['B3'] = date
    
    occ_count = 0
    if ua_ws:
        V = N = kA = kP = sP = leased = 0
        for r in range(8, ua_ws.max_row + 1):
            st = str(ua_ws.cell(r, 1).value or '').strip()
            if st == 'Vacant': V += 1
            if st == 'Notice': N += 1
            if st in ('Vacant', 'Notice'):
                if ua_ws.cell(r, 5).value: kA += 1
                if ua_ws.cell(r, 6).value: kP += 1
                if ua_ws.cell(r, 7).value: sP += 1
            if st == 'Occupied': occ_count += 1
            if st in ('Occupied', 'Notice'):
                try: leased += float(ua_ws.cell(r, 9).value or 0)
                except: pass
        ws['B6'] = -V; ws['B7'] = kA; ws['B8'] = kP; ws['B9'] = sP; ws['B10'] = -N
        ws.cell(16, 2).value = occ_count
        if rr_ws:
            rr_leased = 0
            for rr_r in range(6, rr_ws.max_row + 1):
                rr_name = str(rr_ws.cell(rr_r, 4).value or '').strip()
                if 'VACANT' in rr_name.upper() or not rr_name: continue
                if rr_ws.cell(rr_r, 1).value is None: continue
                try: rr_leased += float(rr_ws.cell(rr_r, 10).value or 0)
                except: pass
            ws['C16'] = rr_leased
        else:
            ws['C16'] = leased
    
    if tar_ws:
        ev = f14 = 0
        for r in range(7, tar_ws.max_row + 1):
            if str(tar_ws.cell(r, 1).value or '').lower() == 'total': break
            st = str(tar_ws.cell(r, 3).value or '').strip().lower()
            if st in ('eviction', 'past'): ev += 1; f14 += 1
        # B14 = count of red-highlighted tenants (any F/G/H/I col >= full rent)
        # This number is computed inside fmt_ar and passed in as tar_red_count
        ws['B11'] = -ev; ws['B14'] = tar_red_count; ws['F14'] = f14
    
    def getT(aw):
        if not aw: return 0
        for r in range(7, aw.max_row + 1):
            if str(aw.cell(r, 1).value or '').lower() == 'total':
                try: return float(aw.cell(r, 5).value or 0)
                except: return 0
        return 0
    if tar_total == 0 and tar_ws: tar_total = getT(tar_ws)
    if sar_total == 0 and sar_ws: sar_total = getT(sar_ws)
    ws['C18'] = tar_total
    ws['C19'] = sar_total
    
    # =========================================================================
    # Column widths
    # =========================================================================
    # Reputation block sits at E:K. Column E is shared with Weekly Traffic
    # 'Source' column (set wide for "Property Website"). Columns F-K need to
    # accommodate the Reputation per-platform table headers.
    widths = {
        'A': 16.57,   # 121px
        'B': 14.43,   # 106px
        'C': 15.43,   # 113px
        'D': 2.71,    # 24px — narrow gutter between B-C and E
        'E': 24.14,   # 174px — wide for Weekly Traffic Source / "ApartmentRatings.com"
        'F': 13.00,   # Average Rating
        'G': 13.00,   # All Time Rating
        'H': 14.00,   # Positive Reviews
        'I': 14.00,   # Negative Reviews
        'J': 13.00,   # Total Reviews / Applications
        'K': 18.00,   # Notes column (Reputation)
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_manager_summary(wb_out, date, prop, ws_summary, traffic_data=None, expiring_rows=None, reputation_data=None):
    """
    Build a manager-friendly bullet-point summary tab using values already computed
    on the Weekly Summary tab. Last tab in workbook.
    """
    BLUE = 'FFBDD7EE'
    tab_name = f'Manager Summary {date}'
    if tab_name in wb_out.sheetnames:
        del wb_out[tab_name]
    ws = wb_out.create_sheet(tab_name)
    
    f10 = gfont(sz=10)
    f10b = gfont(bold=True, sz=10)
    f12b = gfont(bold=True, sz=12)
    f14b = gfont(bold=True, sz=14)
    AB = bblack()
    
    def get_val(coord):
        v = ws_summary[coord].value
        if v is None: return 0
        if isinstance(v, (int, float)): return v
        try: return float(v)
        except: return 0
    
    def fmt_num(n):
        try: return f"{int(n):,}" if abs(n - int(n)) < 0.01 else f"{n:,.2f}"
        except: return str(n)
    
    def fmt_pct(num, denom):
        if not denom: return "0.00%"
        return f"{(num / denom * 100):.2f}%"
    
    def fmt_money(n):
        try: return f"${float(n):,.2f}"
        except: return f"${n}"
    
    # Pull values from Weekly Summary
    total_units = get_val('B5')
    physically_vacant = abs(get_val('B6'))
    kg_approved = get_val('B7')
    kg_pending = get_val('B8')
    site_processing = get_val('B9')
    ntv_not_legal = abs(get_val('B10'))
    ntv_at_legal = abs(get_val('B11'))
    net_leased = get_val('B12')
    tenants_unpaid_full = get_val('B14')
    at_legal = get_val('F14')
    physically_occupied = get_val('B16')
    total_leased_rent = get_val('C16')
    tenant_ar = get_val('C18')
    subsidy_ar = get_val('C19')
    total_ar = tenant_ar + subsidy_ar
    
    # Title block (rows 1-3): merged A:E, bold blue centered
    for r in range(1, 4):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        for c in range(1, 6):
            cell = ws.cell(r, c)
            cell.fill = gfill(BLUE)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = AB
    ws.cell(1, 1).value = prop.split('(')[0].strip()
    ws.cell(1, 1).font = f14b
    ws.cell(2, 1).value = 'Manager Summary'
    ws.cell(2, 1).font = f12b
    ws.cell(3, 1).value = date
    ws.cell(3, 1).font = f10b
    
    # Section header helper
    def section_header(row, text):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row, 1)
        cell.value = text
        cell.font = f12b
        cell.fill = gfill(BLUE)
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        for c in range(1, 6):
            ws.cell(row, c).border = AB
    
    # Bullet helper
    def bullet(row, text):
        cell = ws.cell(row, 1)
        cell.value = f"  •  {text}"
        cell.font = f10
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    
    def ask_questions(row):
        """Add an italic 'Any questions here?' prompt at the end of each category."""
        cell = ws.cell(row, 1)
        cell.value = "Any questions here?"
        cell.font = Font(name='Calibri', size=10, italic=True, color='FF595959')
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=2)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    
    r = 5
    
    # === INTRO LINE (verbal-style summary for the manager to read off) ===
    prop_short = prop.split('(')[0].strip()
    leased_pct = fmt_pct(net_leased, total_units)
    occupied_pct = fmt_pct(physically_occupied, total_units)
    intro_text = f"As for {prop_short} we are currently at {leased_pct} leased, {occupied_pct} occupied"
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    intro_cell = ws.cell(r, 1)
    intro_cell.value = intro_text
    intro_cell.font = gfont(bold=True, sz=11)
    intro_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    r += 1
    r += 1  # blank line before first section
    
    # === OCCUPANCY SECTION ===
    section_header(r, 'AS FOR OCCUPANCY')
    r += 1
    bullet(r, f"Total Units: {fmt_num(total_units)}"); r += 1
    bullet(r, f"Physically Occupied: {fmt_num(physically_occupied)} units ({fmt_pct(physically_occupied, total_units)})"); r += 1
    bullet(r, f"Physically Vacant: {fmt_num(physically_vacant)} units ({fmt_pct(physically_vacant, total_units)})"); r += 1
    bullet(r, f"Net Leased: {fmt_num(net_leased)} units ({fmt_pct(net_leased, total_units)})"); r += 1
    ask_questions(r); r += 1
    r += 1
    
    # === APPLICATIONS / PIPELINE ===
    section_header(r, 'AS FOR THE APPLICATIONS PIPELINE')
    r += 1
    bullet(r, f"Approved @ KG: {fmt_num(kg_approved)}"); r += 1
    bullet(r, f"Pending Not Approved @ KG: {fmt_num(kg_pending)}"); r += 1
    bullet(r, f"Site Processing - Not Sent to KG: {fmt_num(site_processing)}"); r += 1
    ask_questions(r); r += 1
    r += 1
    
    # === NOTICES TO VACATE ===
    section_header(r, 'NOTICES TO VACATE')
    r += 1
    bullet(r, f"Notices to Vacate Not at Legal: {fmt_num(ntv_not_legal)}"); r += 1
    bullet(r, f"Notices to Vacate @ Legal: {fmt_num(ntv_at_legal)}"); r += 1
    ask_questions(r); r += 1
    r += 1
    
    # === DELINQUENCY ===
    section_header(r, 'AS FOR DELINQUENCY')
    r += 1
    bullet(r, f"Tenants who haven't paid full rent: {fmt_num(tenants_unpaid_full)}"); r += 1
    bullet(r, f"Tenants @ legal (eviction): {fmt_num(at_legal)}"); r += 1
    bullet(r, f"Total Leased Rent: {fmt_money(total_leased_rent)}"); r += 1
    bullet(r, f"Tenant Accounts Receivable (AR): {fmt_money(tenant_ar)} ({fmt_pct(tenant_ar, total_leased_rent)})"); r += 1
    bullet(r, f"Subsidy Accounts Receivable (AR): {fmt_money(subsidy_ar)} ({fmt_pct(subsidy_ar, total_leased_rent)})"); r += 1
    bullet(r, f"Total AR: {fmt_money(total_ar)} ({fmt_pct(total_ar, total_leased_rent)})"); r += 1
    ask_questions(r); r += 1
    r += 1
    
    # === WEEKLY TRAFFIC SECTION ===
    if traffic_data and traffic_data.get('rows'):
        date_range = traffic_data.get('date_range', '')
        title = f'AS FOR WEEKLY TRAFFIC ({date_range})' if date_range else 'AS FOR WEEKLY TRAFFIC'
        section_header(r, title); r += 1
        # Tally totals across all sources
        active_sources = [(s, v) for s, v in traffic_data['rows'] if any(x != 0 for x in v)]
        total_leads = sum(v[0] for s, v in active_sources)
        total_pros = sum(v[1] for s, v in active_sources)
        total_visits = sum(v[2] for s, v in active_sources)
        total_leases = sum(v[3] for s, v in active_sources)
        total_apps = sum(v[4] for s, v in active_sources)
        bullet(r, f"Total Leads: {fmt_num(total_leads)}"); r += 1
        bullet(r, f"Total Prospects: {fmt_num(total_pros)}"); r += 1
        bullet(r, f"Total Visits: {fmt_num(total_visits)}"); r += 1
        bullet(r, f"Total Leases Signed: {fmt_num(total_leases)}"); r += 1
        bullet(r, f"Total Applications: {fmt_num(total_apps)}"); r += 1
        # Breakdown by source
        if active_sources:
            r += 1
            for src, vals in active_sources:
                bullet(r, f"{src}: {fmt_num(vals[0])} leads, {fmt_num(vals[1])} prospects, {fmt_num(vals[2])} visits, {fmt_num(vals[3])} leases, {fmt_num(vals[4])} apps")
                r += 1
        ask_questions(r); r += 1
        r += 1
    
    # === EXPIRING LEASES SECTION (day-bucket view) ===
    if expiring_rows:
        from datetime import datetime as _dt, timedelta as _td
        today = _dt.now().replace(hour=0, minute=0, second=0, microsecond=0)
        past_due = 0
        within_30 = 0
        within_60 = 0
        within_90 = 0
        within_120 = 0
        for row in expiring_rows:
            exp_date = row.get('Lease Expires')
            if exp_date is None or not isinstance(exp_date, _dt):
                continue
            days_until = (exp_date - today).days
            if days_until < 0:
                past_due += 1
            else:
                # Cumulative buckets
                if days_until <= 30: within_30 += 1
                if days_until <= 60: within_60 += 1
                if days_until <= 90: within_90 += 1
                if days_until <= 120: within_120 += 1
        total_expiring = past_due + within_120  # past due + everything in next 120 days
        section_header(r, 'AND FOR EXPIRING LEASES'); r += 1
        bullet(r, f"Past due (already expired): {fmt_num(past_due)}"); r += 1
        bullet(r, f"Expiring within 30 days: {fmt_num(within_30)}"); r += 1
        bullet(r, f"Expiring within 60 days: {fmt_num(within_60)}"); r += 1
        bullet(r, f"Expiring within 90 days: {fmt_num(within_90)}"); r += 1
        bullet(r, f"Expiring within 120 days: {fmt_num(within_120)}"); r += 1
        ask_questions(r); r += 1
        r += 1
    
    # === REPUTATION SECTION (Opinionn) — new in v9.50 ===
    if reputation_data:
        section_header(r, 'AS FOR REPUTATION (OPINIONN)'); r += 1
        overall_rating = reputation_data.get('overall_rating') or 0
        overall_total = reputation_data.get('overall_total') or 0
        bullet(r, f"Overall Public Rating: {overall_rating:.1f} stars across {fmt_num(overall_total)} total reviews"); r += 1
        bullet(r, f"Industry Average: 3.9 stars"); r += 1
        # Active platforms (where total > 0)
        active_plats = [(p, d) for p, d in reputation_data.get('platforms', {}).items()
                        if d.get('total', 0) > 0]
        if active_plats:
            r += 1
            for plat, d in active_plats:
                avg = d.get('avg_rating') or 0
                tot = d.get('total') or 0
                bullet(r, f"{plat}: {avg:.1f} stars, {fmt_num(tot)} reviews"); r += 1
        ask_questions(r); r += 1
        r += 1
    
    # Footnote
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    foot = ws.cell(r, 1)
    foot.value = "* AR includes current month delinquency beginning the 10th of each month"
    foot.font = gfont(sz=9)
    foot.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    
    # Column widths
    ws.column_dimensions['A'].width = 80
    for col in 'BCDE':
        ws.column_dimensions[col].width = 12


@app.route('/health')
def health():
    return jsonify({'status':'ok','version':'9.50'})

@app.route('/')
def index():
    return render_template_string(PAGE)

@app.route('/format', methods=['POST'])
def format_report():
    try:
        date=request.form.get('date','').strip()
        prop=request.form.get('prop','').strip()
        if not date or not prop:
            return jsonify({'error':'Missing date or property'}),400
        wb_file=request.files.get('wb')
        if not wb_file:
            return jsonify({'error':'Working workbook is required'}),400
        wb_bytes=wb_file.read()
        wb_ro=openpyxl.load_workbook(io.BytesIO(wb_bytes),data_only=True,keep_vba=False,read_only=True)
        pTAR=get_notes(wb_ro,'Tenant AR')
        pSAR=get_notes(wb_ro,'Sub AR')
        pSAR.update(get_notes(wb_ro,'SUB AR'))
        wb_out=openpyxl.Workbook(); wb_out.remove(wb_out.active)
        ua_ws=tar_ws=sar_ws=None
        tar_total=sar_total=0
        ua_f=request.files.get('ua')
        if ua_f: ua_ws,*_=fmt_ua(wb_out,ua_f.read(),date,prop)
        # Process Rent Roll FIRST so we can build a unit->full_rent lookup for Tenant AR highlighting
        rr_ws=None
        rent_lookup={}
        rr_f=request.files.get('rr')
        if rr_f: rr_ws,*_=fmt_rr(wb_out,rr_f.read(),date,prop)
        if rr_ws:
            # Build {unit: tenant_rent} dict from rent roll (col 1 = Unit, col 9 = Tenant Rent)
            for rr_r in range(6, rr_ws.max_row + 1):
                unit_val = str(rr_ws.cell(rr_r, 1).value or '').strip()
                if not unit_val: continue
                try:
                    tenant_rent_val = float(rr_ws.cell(rr_r, 9).value or 0)
                    if tenant_rent_val > 0:
                        rent_lookup[unit_val] = tenant_rent_val
                except: pass
        tar_f=request.files.get('tar')
        tar_red_count=0
        if tar_f:
            tar_ws,ev,cu,no,pos_end,data_start,tar_red_count=fmt_ar(wb_out,tar_f.read(),date,pTAR,False,rent_lookup)
            for r in range(data_start,pos_end+1):
                try: tar_total+=float(tar_ws.cell(r,5).value or 0)
                except: pass
        sar_f=request.files.get('sar')
        if sar_f:
            sar_ws,ev,cu,no,pos_end,data_start,_=fmt_ar(wb_out,sar_f.read(),date,pSAR,True)
            for r in range(data_start,pos_end+1):
                try: sar_total+=float(sar_ws.cell(r,5).value or 0)
                except: pass
        traffic_data=None
        tr_f=request.files.get('tr')
        if tr_f: traffic_data=parse_traffic(tr_f.read(),date)
        # Expiring Leases
        expiring_rows=None
        ex_f=request.files.get('ex')
        if ex_f:
            ex_bytes = ex_f.read()
            try:
                _, expiring_rows = fmt_expiring(wb_out, ex_bytes, date, prop)
            except Exception:
                traceback.print_exc()
                expiring_rows = None
        # Opinionn Review Summary PDF (new in v9.50)
        reputation_data=None
        op_f=request.files.get('op')
        if op_f:
            try:
                op_bytes = op_f.read()
                reputation_data = parse_review_summary_pdf(op_bytes)
            except Exception:
                traceback.print_exc()
                reputation_data = None
        build_weekly_summary(wb_out,wb_ro,date,prop,ua_ws,tar_ws,sar_ws,tar_total,sar_total,rr_ws,traffic_data,expiring_rows,tar_red_count,reputation_data)
        # Build Manager Summary tab (reads computed values from Weekly Summary)
        ws_summary_name = next((n for n in wb_out.sheetnames if 'weekly summary' in n.lower()), None)
        if ws_summary_name:
            build_manager_summary(wb_out, date, prop, wb_out[ws_summary_name], traffic_data, expiring_rows, reputation_data)
        wb_ro.close()
        def find_tab(names, prefix):
            return next((n for n in names if n.strip().lower().startswith(prefix.lower())), None)
        current=list(wb_out.sheetnames)
        ws_tab=find_tab(current,'weekly summary'); ua_tab=find_tab(current,'unit availability')
        rr_tab=find_tab(current,'rent roll'); tar_tab=find_tab(current,'tenant ar'); sar_tab=find_tab(current,'sub ar')
        ex_tab=find_tab(current,'expiring leases'); mgr_tab=find_tab(current,'manager summary')
        desired=[t for t in [ws_tab,ua_tab,rr_tab,tar_tab,sar_tab,ex_tab,mgr_tab] if t]
        remaining=[t for t in current if t not in desired]
        ordered=desired+remaining
        for target_i,name in enumerate(ordered):
            current=list(wb_out.sheetnames); current_i=current.index(name)
            if current_i!=target_i: wb_out.move_sheet(name,offset=target_i-current_i)
        import zipfile, re as _re
        raw=io.BytesIO(); wb_out.save(raw); raw.seek(0)
        patched=io.BytesIO()
        with zipfile.ZipFile(raw,'r') as zin, zipfile.ZipFile(patched,'w',zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                data=zin.read(item)
                if item=='xl/workbook.xml':
                    txt=data.decode('utf-8')
                    txt=_re.sub(r'<calcPr[^/]*/>',"<calcPr calcId=\"191\" refMode=\"A1\"/>",txt)
                    data=txt.encode('utf-8')
                zout.writestr(item,data)
        patched.seek(0)
        prefix=prop.split('(')[0].strip().replace(' ','_')
        fname=f'{prefix}_Weekly_{date.replace(".","")}_Formatted.xlsx'
        return send_file(patched,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',as_attachment=True,download_name=fname)
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error':str(e)}),500

PAGE='''<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Weekly Report Formatter</title>
<style>
@import url("https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=Sora:wght@300;400;600;700&display=swap");
:root{--g:#7AD694;--gd:#4fa868;--bg:#f5f6f8;--card:#fff;--bdr:#e0e2e8;--ink:#1a1c24;--mut:#7a7e90;}
*{box-sizing:border-box;margin:0;padding:0;}
body{font-family:"Sora",sans-serif;background:var(--bg);color:var(--ink);}
.hdr{background:linear-gradient(135deg,#1a2e1e,#2c4a31);padding:24px 40px;display:flex;align-items:center;gap:16px;border-bottom:3px solid var(--g);}
.hi{width:40px;height:40px;background:var(--g);border-radius:9px;display:flex;align-items:center;justify-content:center;font-size:20px;}
.hdr h1{font-size:18px;font-weight:700;color:#fff;}
.hdr p{font-size:11px;color:rgba(255,255,255,.45);font-family:"DM Mono",monospace;margin-top:2px;}
.hv{margin-left:auto;background:rgba(122,214,148,.15);border:1px solid rgba(122,214,148,.3);color:var(--g);padding:4px 11px;border-radius:20px;font-size:11px;font-family:"DM Mono",monospace;}
.main{max-width:800px;margin:0 auto;padding:28px 20px 60px;}
.card{background:var(--card);border:1px solid var(--bdr);border-radius:13px;padding:20px 24px;margin-bottom:14px;position:relative;}
.sn{position:absolute;top:-1px;left:20px;background:var(--g);color:#1a2e1e;font-size:10px;font-weight:700;font-family:"DM Mono",monospace;padding:3px 10px;border-radius:0 0 7px 7px;letter-spacing:1px;}
.ct{font-size:14px;font-weight:700;margin:6px 0 4px;}
.cd{font-size:12px;color:var(--mut);margin-bottom:14px;line-height:1.6;}
select,input[type=text]{font-family:"DM Mono",monospace;font-size:13px;border:1.5px solid var(--bdr);border-radius:8px;padding:8px 12px;background:#fff;color:var(--ink);transition:border-color .2s;outline:none;}
select:focus,input:focus{border-color:var(--g);}
.grid{display:grid;grid-template-columns:1fr 1fr;gap:10px;}
.slot{border:1.5px dashed var(--bdr);border-radius:10px;padding:12px 14px;background:#fafbfc;position:relative;transition:all .2s;cursor:pointer;}
.slot:hover{border-color:var(--g);background:rgba(122,214,148,.04);}
.slot.on{border-style:solid;border-color:var(--gd);background:#f0fbf3;}
.slot.full{grid-column:1/-1;}
.slot input[type=file]{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%;}
.sh{display:flex;align-items:center;gap:7px;margin-bottom:3px;}
.dot{width:8px;height:8px;border-radius:50%;flex-shrink:0;}
.sl{font-size:12px;font-weight:700;}
.ss{font-size:10.5px;color:var(--mut);line-height:1.4;}
.sn2{font-size:10.5px;color:var(--gd);font-family:"DM Mono",monospace;margin-top:4px;min-height:14px;}
.btn{display:block;width:100%;background:linear-gradient(135deg,#2c4a31,#1f3824);color:#fff;border:none;border-radius:10px;padding:14px;font-family:"Sora",sans-serif;font-size:14px;font-weight:700;cursor:pointer;transition:all .2s;margin-top:6px;}
.btn:hover:not(:disabled){transform:translateY(-1px);box-shadow:0 8px 24px rgba(44,74,49,.35);}
.btn:disabled{opacity:.5;cursor:not-allowed;}
.pg{height:4px;background:var(--bdr);border-radius:2px;overflow:hidden;margin-top:10px;display:none;}
.pg.on{display:block;}
.pf{height:100%;background:var(--g);border-radius:2px;transition:width .4s;}
.log{background:#1a1c24;border-radius:9px;padding:12px;font-family:"DM Mono",monospace;font-size:11px;color:#8aff9a;max-height:180px;overflow-y:auto;display:none;line-height:1.7;margin-top:10px;}
.log.on{display:block;}
.le{color:#f28e86;}.lw{color:#fdd868;}.li{color:#8cb5f9;}
.dl{display:none;background:#f0fbf3;border:2px solid var(--g);border-radius:11px;padding:18px;text-align:center;margin-top:12px;}
.dl.on{display:block;}
.dl h3{font-size:14px;font-weight:700;color:var(--gd);margin-bottom:5px;}
.dl p{font-size:12px;color:var(--mut);margin-bottom:12px;}
.dlb{display:inline-flex;align-items:center;gap:7px;background:var(--gd);color:#fff;border:none;border-radius:8px;padding:10px 22px;font-family:"Sora",sans-serif;font-size:13px;font-weight:700;cursor:pointer;text-decoration:none;}
.dlb:hover{background:#3d8a53;}
@media(max-width:600px){.hdr{padding:16px;}.main{padding:16px 12px 50px;}.grid{grid-template-columns:1fr;}.slot.full{grid-column:1;}}
</style></head><body>
<div class="hdr"><div class="hi">&#127970;</div><div><h1>Weekly Report Formatter</h1><p>Occupancy &amp; Delinquency &middot; FPI Management</p></div><div class="hv">v9.50</div></div>
<div class="main">
  <div class="card"><div class="sn">STEP 01</div><div class="ct">Select Property &amp; Enter Date</div><div class="cd">Choose the property and enter this week\'s report date.</div>
    <select id="prop" style="width:100%;margin-bottom:10px;"><option value="Village at Madrone (fka Village at Morgan Hill) (x93)">Village at Madrone (x93)</option><option value="Village at First">Village at First</option><option value="Village at Santa Teresa">Village at Santa Teresa</option></select>
    <div style="display:flex;align-items:center;gap:10px;"><input type="text" id="date" placeholder="04.06.26" maxlength="8" style="width:120px;"/><span style="font-size:11px;color:var(--mut);font-family:\'DM Mono\',monospace;">MM.DD.YY</span></div>
  </div>
  <div class="card"><div class="sn">STEP 02</div><div class="ct">Upload Working Workbook</div><div class="cd">The master workbook with Weekly Summary and prior AR history.</div>
    <div class="grid"><div class="slot full" id="s-wb"><input type="file" id="f-wb" accept=".xlsx,.xls,.xlsm"/><div class="sh"><div class="dot" style="background:#8CB5F9;"></div><span class="sl">&#128210; Weekly Workbook</span></div><div class="ss">Master file — Weekly Summary + history</div><div class="sn2" id="n-wb">Click or drag file here</div></div></div>
  </div>
  <div class="card"><div class="sn">STEP 03</div><div class="ct">Upload Yardi Exports &amp; Opinionn PDF</div><div class="cd">Upload each Yardi export. Leave empty anything you don\'t have — it will be skipped.</div>
    <div class="grid">
      <div class="slot" id="s-ua"><input type="file" id="f-ua" accept=".xlsx,.xls,.xlsm"/><div class="sh"><div class="dot" style="background:#7AD694;"></div><span class="sl">Unit Availability</span></div><div class="ss">Onsite &rarr; Analytics &rarr; Unit Availability Details</div><div class="sn2" id="n-ua">Click or drag file here</div></div>
      <div class="slot" id="s-tar"><input type="file" id="f-tar" accept=".xlsx,.xls,.xlsm"/><div class="sh"><div class="dot" style="background:#F28E86;"></div><span class="sl">Tenant AR</span></div><div class="ss">Analytics &rarr; Receivable Aging (Excl. HUD)</div><div class="sn2" id="n-tar">Click or drag file here</div></div>
      <div class="slot" id="s-sar"><input type="file" id="f-sar" accept=".xlsx,.xls,.xlsm"/><div class="sh"><div class="dot" style="background:#8CB5F9;"></div><span class="sl">Subsidy AR</span></div><div class="ss">Analytics &rarr; Receivable Aging (HUD Only)</div><div class="sn2" id="n-sar">Click or drag file here</div></div>
      <div class="slot" id="s-rr"><input type="file" id="f-rr" accept=".xlsx,.xls,.xlsm"/><div class="sh"><div class="dot" style="background:#C4A0F5;"></div><span class="sl">Rent Roll</span></div><div class="ss">Onsite &rarr; Analytics &rarr; Rent Roll</div><div class="sn2" id="n-rr">Click or drag file here</div></div>
      <div class="slot" id="s-tr"><input type="file" id="f-tr" accept=".xlsx,.xls,.xlsm,.csv"/><div class="sh"><div class="dot" style="background:#F5C842;"></div><span class="sl">Weekly Traffic</span></div><div class="ss">Ad Spend &amp; Traffic Report (CSV or Excel)</div><div class="sn2" id="n-tr">Click or drag file here</div></div>
      <div class="slot" id="s-ex"><input type="file" id="f-ex" accept=".xlsx,.xls,.xlsm"/><div class="sh"><div class="dot" style="background:#FDD868;"></div><span class="sl">Expiring Leases (120 days)</span></div><div class="ss">Analytics &rarr; Expiring Leases</div><div class="sn2" id="n-ex">Click or drag file here</div></div>
      <div class="slot full" id="s-op"><input type="file" id="f-op" accept=".pdf"/><div class="sh"><div class="dot" style="background:#2DD4BF;"></div><span class="sl">Opinionn Review Summary (PDF)</span></div><div class="ss">Opinionn &rarr; Reports &rarr; Review Summary &rarr; download PDF</div><div class="sn2" id="n-op">Click or drag PDF here</div></div>
    </div>
  </div>
  <div class="card"><div class="sn">STEP 04</div><div class="ct">Format &amp; Download</div><div class="cd">Formats all reports with exact colors, structure, and formulas. Downloads the final workbook.</div>
    <button class="btn" id="btn" onclick="run()">&#9889; Format Report</button>
    <div class="pg" id="pg"><div class="pf" id="pf" style="width:0%"></div></div>
    <div class="log" id="log"></div>
    <div class="dl" id="dl"><h3>&#10003; Done!</h3><p>Your formatted workbook is ready.</p><a class="dlb" id="dlb" href="#">&#8595; Download Formatted Workbook</a></div>
  </div>
</div>
<script>
["wb","ua","tar","sar","rr","tr","ex","op"].forEach(k=>{
  document.getElementById("f-"+k).addEventListener("change",function(){
    if(this.files[0]){document.getElementById("s-"+k).classList.add("on");document.getElementById("n-"+k).textContent="✓ "+this.files[0].name;}
  });
});
function L(m,c=""){const el=document.getElementById("log");el.classList.add("on");const d=document.createElement("div");if(c)d.className=c;d.textContent="> "+m;el.appendChild(d);el.scrollTop=el.scrollHeight;}
function P(p){document.getElementById("pg").classList.add("on");document.getElementById("pf").style.width=p+"%";}
async function run(){
  const btn=document.getElementById("btn");btn.disabled=true;
  document.getElementById("log").innerHTML="";document.getElementById("log").classList.remove("on");
  document.getElementById("dl").classList.remove("on");document.getElementById("pg").classList.remove("on");
  const date=document.getElementById("date").value.trim();const prop=document.getElementById("prop").value;
  if(!date){alert("Please enter the report date.");btn.disabled=false;return;}
  if(!document.getElementById("f-wb").files[0]){alert("Please upload the working workbook.");btn.disabled=false;return;}
  const form=new FormData();
  form.append("date",date);form.append("prop",prop);
  ["wb","ua","tar","sar","rr","tr","ex","op"].forEach(k=>{const f=document.getElementById("f-"+k).files[0];if(f)form.append(k,f);});
  L("Uploading and formatting...");P(20);
  try{
    const resp=await fetch("/format",{method:"POST",body:form});
    P(80);
    if(!resp.ok){
      let msg="Server error";
      try{const e=await resp.json();msg=e.error||msg;}catch(ex){}
      L("Error: "+msg,"le");btn.disabled=false;return;
    }
    const blob=await resp.blob();P(100);
    L("Done!","li");
    const url=URL.createObjectURL(blob);
    const prefix=prop.split("(")[0].trim().replace(/ /g,"_");
    const fname=prefix+"_Weekly_"+date.replace(/\\./g,"")+"_Formatted.xlsx";
    const a=document.getElementById("dlb");a.href=url;a.download=fname;
    document.getElementById("dl").classList.add("on");
  }catch(e){L("Error: "+e.message,"le");}
  btn.disabled=false;
}
</script></body></html>'''

if __name__=='__main__':
    port=int(os.environ.get('PORT',5000))
    app.run(host='0.0.0.0',port=port,debug=False)
